"""
Импорт записей в журнал проверки знаний из Excel файла
"""
import os
import sys
import uuid
from datetime import datetime
from openpyxl import load_workbook

project_dir = os.path.dirname(os.path.abspath(__file__))
data_dir = os.path.join(project_dir, "data")

from journal.journal_manager import JournalManager
from db.exam_journal_repo import JournalRecord
from utils.crypto import encrypt_data, decrypt_data

def import_journal_from_excel(file_path: str) -> tuple[int, list]:
    """
    Импорт записей из Excel файла в журнал
    
    Колонки Excel:
    0: Фамилия
    1: Имя  
    2: Отчество
    3: СНИЛС
    4: Должность
    5: № программы
    6: Название программы
    7: Дата экзамена
    8: № протокола
    9: Результат
    10: SetId
    11: Рег. номер (опционально)
    
    Возвращает (количество_добавленных, список_ошибок)
    """
    if not os.path.exists(file_path):
        return 0, [f"Файл не найден: {file_path}"]
    
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        # Читаем заголовки
        headers = [str(ws.cell(row=1, column=c).value).strip() for c in range(1, ws.max_column + 1)]
        
        # Проверяем обязательные колонки
        required = ['Фамилия', 'Имя', 'СНИЛС', '№ программы', 'Дата экзамена', '№ протокола', 'Результат']
        missing = [col for col in required if col not in headers]
        if missing:
            return 0, [f"Отсутствуют колонки: {', '.join(missing)}"]
        
        # Инициализируем журнал
        journal = JournalManager(data_dir)
        
        added_count = 0
        errors = []
        records_to_add = []
        
        for row_num in range(2, ws.max_row + 1):
            try:
                # Читаем строку
                row_dict = {}
                for c_idx, col_name in enumerate(headers):
                    val = ws.cell(row=row_num, column=c_idx + 1).value
                    row_dict[col_name] = val if val is not None else ''
                
                # Пропускаем пустые строки
                if not row_dict.get('Фамилия') or not row_dict.get('СНИЛС'):
                    continue
                
                # Форматируем СНИЛС
                snils_raw = str(row_dict.get('СНИЛС', '')).strip()
                snils = format_snils(snils_raw)
                if not snils:
                    errors.append(f"Строка {row_num}: некорректный СНИЛС")
                    continue
                
                # Создаем запись
                record = JournalRecord(
                    uuid=str(uuid.uuid4()),
                    send_date=datetime.now().strftime("%d.%m.%Y %H:%M:%S"),
                    set_id=str(row_dict.get('SetId', '')).strip() or f"IMPORT-{datetime.now().strftime('%Y%m%d')}",
                    xml_file="",
                    last_name=str(row_dict.get('Фамилия', '')).strip(),
                    first_name=str(row_dict.get('Имя', '')).strip(),
                    middle_name=str(row_dict.get('Отчество', '')).strip(),
                    snils=snils,
                    position=str(row_dict.get('Должность', '')).strip(),
                    program_id=str(row_dict.get('№ программы', '')).strip(),
                    program_title=str(row_dict.get('Название программы', '')).strip() or f"Программа {row_dict.get('№ программы', '')}",
                    exam_date=str(row_dict.get('Дата экзамена', '')).strip(),
                    protocol=str(row_dict.get('№ протокола', '')).strip(),
                    result=str(row_dict.get('Результат', 'Удовлетворительно')).strip(),
                    base_no=str(row_dict.get('Рег. номер', '')).strip(),
                    status="pending" if not row_dict.get('Рег. номер') else "received"
                )
                
                records_to_add.append(record)
                added_count += 1
                
            except Exception as e:
                errors.append(f"Строка {row_num}: {str(e)}")
        
        # Сохраняем журнал
        if added_count > 0:
            journal.add_journal_records_directly(records_to_add)
        
        return added_count, errors
        
    except Exception as e:
        return 0, [f"Ошибка чтения файла: {str(e)}"]


def format_snils(raw: str) -> str:
    """Форматирование СНИЛС в вид '123-456-789 00'"""
    import unicodedata
    clean = ''.join(c for c in str(raw) if unicodedata.category(c) != 'Zs')
    clean = clean.replace('-', '').replace(' ', '')
    if not clean.isdigit() or len(clean) != 11:
        return None
    return f"{clean[0:3]}-{clean[3:6]}-{clean[6:9]} {clean[9:11]}"


if __name__ == "__main__":
    # Путь к тестовому файлу (создайте test_data/journal_import.xlsx для теста)
    test_dir = os.path.join(project_dir, "test_data")
    file_path = os.path.join(test_dir, "journal_import.xlsx")
    if not os.path.exists(test_dir):
        os.makedirs(test_dir, exist_ok=True)
    
    print("=== Импорт в журнал ===")
    print(f"Файл: {file_path}")
    print(f"Существует: {os.path.exists(file_path)}")
    
    count, errors = import_journal_from_excel(file_path)
    
    print(f"\nДобавлено записей: {count}")
    if errors:
        print(f"Ошибок: {len(errors)}")
        for err in errors[:5]:
            print(f"  - {err}")
    else:
        print("Ошибок нет!")
    
    # Проверяем журнал
    journal = JournalManager(data_dir)
    print(f"\nВсего записей в журнале: {journal.get_record_count()}")