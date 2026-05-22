import os
import json
import time
import logging
from datetime import datetime
from typing import List
from dateutil.relativedelta import relativedelta

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGroupBox, QFormLayout,
    QLabel, QLineEdit, QPushButton, QComboBox, QScrollArea,
    QFileDialog, QMessageBox, QTableWidget, QTableWidgetItem,
    QHeaderView, QAbstractItemView, QFrame, QCheckBox,
    QListWidget, QListWidgetItem, QMenu, QDialog
)
from utils.error_utils import safe_message_box
from utils.dialog_base import BaseDialog
from utils.field_validators import ValidatedLineEdit
from utils.constants import VALID_PROGRAMS, VALID_PROGRAMS_SET, DEFAULT_PROGRAMS, PROGRAM_TITLES
from PySide6.QtCore import Qt, QThread, Signal
from PySide6.QtGui import QColor, QFont, QPalette

from db import (
    DatabaseManager, EmployeesRepo, EmployeeProgramsRepo
)
from api.mintrud_api import load_api_key, get_by_snils
from utils.proxy_manager import load_proxy_settings
from utils.training_rules import get_dynamic_status, compute_expiry_date, get_training_period_years
from utils.audit import log_audit
from utils.logger import filter_sensitive_text

logger = logging.getLogger(__name__)


def _normalize_api_date(date_str: str) -> str:
    if not date_str:
        return ""
    date_str = date_str.strip()
    if '.' in date_str:
        parts = date_str.split()[0]
        if len(parts) == 10 and parts[2] == '.' and parts[5] == '.':
            return parts
    if 'T' in date_str:
        date_str = date_str.split('T')[0]
    if '-' in date_str:
        parts = date_str.split()[0]
        try:
            dt = datetime.strptime(parts, "%Y-%m-%d")
            return dt.strftime("%d.%m.%Y")
        except (ValueError, IndexError):
            pass
    return date_str


def _dmy_gt(a: str, b: str) -> bool:
    """Compare two DD.MM.YYYY dates, returns True if a > b chronologically."""
    try:
        return datetime.strptime(a.strip(), '%d.%m.%Y') > datetime.strptime(b.strip(), '%d.%m.%Y')
    except (ValueError, AttributeError):
        return False


BASE_COLUMNS = 3
SUB_COLUMNS = 4


class ApiQueryThread(QThread):
    finished = Signal(int, int)
    error_signal = Signal(str)
    progress = Signal(int, int)

    def __init__(self, employees, api_key, proxy_settings=None):
        super().__init__()
        self.employees = employees
        self.api_key = api_key
        self.proxy_settings = proxy_settings or {}

    def run(self):
        total = len(self.employees)
        updated = 0
        errors = 0
        for idx, emp in enumerate(self.employees):
            try:
                snils_clean = emp['snils'].replace('-', '').replace(' ', '')
                result = get_by_snils(self.api_key, snils_clean, proxy_settings=self.proxy_settings)
                if result.get("success"):
                    records = result.get("records", [])
                    self._process_api_records(emp['id'], records)
                    updated += 1
                    now = datetime.now().strftime('%d.%m.%Y %H:%M:%S')
                    EmployeesRepo.update_sync(emp['id'], now)
                else:
                    errors += 1
                    logger.warning("API error: %s", filter_sensitive_text(str(result.get('error'))))
            except Exception as e:
                errors += 1
                logger.exception("Registry API exception")
            self.progress.emit(idx + 1, total)
            if idx < total - 1:
                time.sleep(0.5)
        DatabaseManager.close_thread_connection()
        self.finished.emit(updated, errors)

    def _process_api_records(self, employee_id, records):
        seen_programs = {}
        for rec in records:
            prog_id = rec.get('learnProgramId', '')
            if not prog_id:
                continue
            exam_date = _normalize_api_date(rec.get('Date', ''))
            protocol = rec.get('ProtocolNumber', '')
            base_no = rec.get('baseNo', '')
            is_passed = rec.get('isPassed', '')
            result = 1 if is_passed and is_passed.lower() in ('true', '1', 'да', 'удовлетворительно') else 0
            if prog_id not in seen_programs or (exam_date and _dmy_gt(exam_date, seen_programs[prog_id].get('exam_date', ''))):
                seen_programs[prog_id] = { 'exam_date': exam_date,
                    'protocol': protocol, 'base_no': base_no, 'result': result,}
        for prog_id, data in seen_programs.items():
            try:
                EmployeeProgramsRepo.update_from_api(
                    employee_id, int(prog_id),
                    data['exam_date'], data['protocol'], data['base_no'], data['result']
                )
            except (ValueError, TypeError):
                pass
        if records:
            first_rec = records[0]
            emp = EmployeesRepo.get_by_id(employee_id)
            if emp:
                updates = {}
                if not emp.get('last_name'): updates['last_name'] = first_rec.get('LastName', '')
                if not emp.get('first_name'): updates['first_name'] = first_rec.get('FirstName', '')
                if not emp.get('middle_name'): updates['middle_name'] = first_rec.get('MiddleName', '')
                if not emp.get('position'): updates['position'] = first_rec.get('Position', '')
                if updates:
                    updates['snils'] = emp['snils']
                    EmployeesRepo.upsert(updates)


class PlanDialog(BaseDialog):
    def __init__(self, plan_data: list, plan_title: str, parent=None):
        super().__init__(parent, title=plan_title, min_width=900, min_height=600)

        bl = self.body_layout()

        stats_layout = QHBoxLayout()
        total = len(plan_data)
        high = sum(1 for p in plan_data if p['priority'] == 'Высокий')
        medium = sum(1 for p in plan_data if p['priority'] == 'Средний')
        low = sum(1 for p in plan_data if p['priority'] == 'Низкий')

        for label, value, color in [
            ("Всего в плане", str(total), "#4169E1"),
            ("Высокий", str(high), "#dc3545"),
            ("Средний", str(medium), "#ffc107"),
            ("Низкий", str(low), "#28a745"),
        ]:
            card = QFrame()
            card.setObjectName("planStatCard")
            card.setStyleSheet(f"""
                QFrame {{ border: 2px solid {color}; border-radius: 8px;
                    padding: 10px;}}
            """)
            card_layout = QVBoxLayout(card)
            val_label = QLabel(value)
            val_label.setStyleSheet(f"font-size: 28px; font-weight: bold; color: {color}; background-color: transparent;")
            val_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            desc_label = QLabel(label)
            desc_label.setStyleSheet("font-size: 12px; background-color: transparent;")
            desc_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            card_layout.addWidget(val_label)
            card_layout.addWidget(desc_label)
            stats_layout.addWidget(card)
        bl.addLayout(stats_layout)

        table = QTableWidget()
        table.setColumnCount(9)
        table.setHorizontalHeaderLabels([
            "№", "ФИО", "СНИЛС", "Должность", "Программа",
            "Последняя\nдата\nобучения", "Дата\nокончания\nдействия",
            "Основание", "Приоритет"
        ])
        table.setAlternatingRowColors(True)
        table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        table.verticalHeader().setVisible(False)
        table.horizontalHeader().setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
        table.setWordWrap(True)

        for i, p in enumerate(plan_data, 1):
            row = table.rowCount()
            table.insertRow(row)
            items = [
                str(i),
                f"{p['last_name']} {p['first_name']} {p['middle_name']}".strip(),
                p['snils'], p['position'], p['program'],
                p['last_exam_date'], p['expiry_date'],
                p['reason'], p['priority'],
            ]
            for col, text in enumerate(items):
                item = QTableWidgetItem(str(text))
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if col == 8:
                    color_map = {'Высокий': '#dc3545', 'Средний': '#ffc107', 'Низкий': '#28a745'}
                    item.setForeground(QColor(color_map.get(text, '#E0E0E0')))
                    font = item.font(); font.setBold(True); item.setFont(font)
                table.setItem(row, col, item)

        bl.addWidget(table)

        btn_layout = QHBoxLayout()
        export_btn = QPushButton("Экспорт XLSX")
        export_btn.setObjectName("planExportBtn")
        export_btn.clicked.connect(lambda: self._export_plan(plan_data, plan_title))
        close_btn = QPushButton("Закрыть")
        close_btn.setObjectName("planCloseBtn")
        close_btn.clicked.connect(self.close)
        btn_layout.addWidget(export_btn)
        btn_layout.addStretch()
        btn_layout.addWidget(close_btn)
        bl.addLayout(btn_layout)

    def _export_plan(self, plan_data: list, plan_title: str):
        from utils.toast import Toast
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить XLSX", f"{plan_title}.xlsx", "Excel Files (*.xlsx)"
        )
        if not file_path:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = Workbook()
            ws = wb.active
            ws.title = "План"
            headers = ["№", "ФИО", "СНИЛС", "Должность", "Программа",
                "Последняя\nдата\nобучения", "Дата\nокончания\nдействия", "Основание", "Приоритет"]
            ws.append(headers)
            hf = Font(bold=True, color="FFFFFF")
            hfill = PatternFill(start_color="4169E1", end_color="4169E1", fill_type="solid")
            for cell in ws[1]:
                cell.font = hf; cell.fill = hfill; cell.alignment = Alignment(horizontal="center")
            for i, p in enumerate(plan_data, 1):
                ws.append([i, f"{p['last_name']} {p['first_name']} {p['middle_name']}".strip(),
                    p['snils'], p['position'], p['program'],
                    p['last_exam_date'], p['expiry_date'], p['reason'], p['priority']])
            ws2 = wb.create_sheet("Сводка")
            ws2.append(["Программа", "Кол-во"])
            from collections import Counter
            for prog, cnt in sorted(Counter(p['program'] for p in plan_data).items()):
                ws2.append([prog, cnt])
            for cell in ws2[1]:
                cell.font = hf; cell.fill = hfill; cell.alignment = Alignment(horizontal="center")
            wb.save(file_path)
            Toast.success(self, f"Файл сохранён:\n{file_path}")
        except Exception as e:
            logger.exception("Export error")
            from utils.error_utils import show_error_dialog
            show_error_dialog(self, "Ошибка экспорта", str(e), critical=False)


# ──────── EmployeeAddDialog ────────

class EmployeeAddDialog(BaseDialog):
    def __init__(self, parent=None):
        super().__init__(parent, title="Добавление сотрудника", min_width=520, min_height=400)

        bl = self.body_layout()
        form = QFormLayout()
        form.setSpacing(8)

        from utils.field_validators import validate_snils, validate_name

        self.last_name = ValidatedLineEdit(validator=validate_name)
        self.last_name.setPlaceholderText("Иванов")
        form.addRow("Фамилия:", self.last_name)

        self.first_name = ValidatedLineEdit(validator=validate_name)
        self.first_name.setPlaceholderText("Иван")
        form.addRow("Имя:", self.first_name)

        self.middle_name = ValidatedLineEdit(validator=validate_name)
        self.middle_name.setPlaceholderText("Иванович")
        form.addRow("Отчество:", self.middle_name)

        self.snils = ValidatedLineEdit(validator=validate_snils)
        self.snils.setPlaceholderText("123-456-789 00")
        self.register_field(self.snils)
        form.addRow("СНИЛС:", self.snils)

        self.position = ValidatedLineEdit()
        self.position.setPlaceholderText("Слесарь")
        form.addRow("Должность:", self.position)

        prog_row = QHBoxLayout()
        self.programs = QLineEdit()
        self.programs.setPlaceholderText("1, 2, 3")
        prog_row.addWidget(self.programs, 1)
        help_btn = QPushButton("Справка")
        help_btn.setObjectName("programHelpBtn")
        help_btn.clicked.connect(self._show_programs_help)
        prog_row.addWidget(help_btn)
        form.addRow("Программы:", prog_row)

        bl.addLayout(form)

        btn_row = QHBoxLayout()
        save_btn = QPushButton("Добавить"); save_btn.setObjectName("dialogPrimaryBtn")
        save_btn.clicked.connect(self._validate_and_accept)
        cancel_btn = QPushButton("Отмена"); cancel_btn.setObjectName("dialogDangerBtn")
        cancel_btn.clicked.connect(self.reject)
        btn_row.addWidget(save_btn); btn_row.addStretch(); btn_row.addWidget(cancel_btn)
        bl.addLayout(btn_row)

    def _validate_and_accept(self):
        from utils.field_validators import validate_program_id
        if not self.validate_fields():
            return
        prog_text = self.programs.text().strip()
        if prog_text:
            pids = [p.strip() for p in prog_text.split(",") if p.strip()]
            for pid in pids:
                err = validate_program_id(pid)
                if err:
                    self.programs.setStyleSheet("border: 2px solid #E74C3C; background-color: #FFF0F0;")
                    self.programs.setToolTip(f"Некорректная программа: {pid}")
                    self.programs.setFocus()
                    return
        self.programs.setStyleSheet("")
        self.accept()

    def _show_programs_help(self):
        from PySide6.QtGui import QPalette
        d = QDialog(self)
        d.setWindowTitle("Программы обучения"); d.setMinimumSize(600, 500)
        dl = QVBoxLayout(d)
        lw = QListWidget(); lw.setSelectionMode(QListWidget.SelectionMode.MultiSelection)
        current = [p.strip() for p in self.programs.text().split(",") if p.strip()]
        blue = {"1", "2", "3", "4", "18", "23"}

        pal = d.palette()
        primary = pal.color(QPalette.ColorRole.Highlight)
        current_bg = QColor(primary.red(), primary.green(), primary.blue(), 35)

        for pid in VALID_PROGRAMS:
            title = PROGRAM_TITLES.get(pid, "")
            item = QListWidgetItem(f"{pid}: {title}")
            item.setData(Qt.ItemDataRole.UserRole, pid)
            if pid in current:
                item.setBackground(current_bg)
            if pid in blue:
                item.setForeground(primary)
            lw.addItem(item)
        def add_sel():
            sel = lw.selectedItems()
            if not sel: return
            new_set = set(current)
            for it in sel: new_set.add(it.data(Qt.ItemDataRole.UserRole))
            self.programs.setText(", ".join(sorted(new_set, key=lambda x: int(x) if x.isdigit() else 0)))
            d.accept()
        lw.itemDoubleClicked.connect(lambda item: (
            current.append(item.data(Qt.ItemDataRole.UserRole))
            or self.programs.setText(", ".join(current))))
        dl.addWidget(lw)
        br = QHBoxLayout()
        ab = QPushButton("Добавить выбранные"); ab.setObjectName("dialogPrimaryBtn"); ab.clicked.connect(add_sel)
        br.addWidget(ab); br.addStretch()
        cb = QPushButton("Отмена"); cb.setObjectName("dialogDangerBtn"); cb.clicked.connect(d.reject)
        br.addWidget(cb); dl.addLayout(br)
        d.exec()


# ──────── EmployeeEditDialog ────────

class EmployeeEditDialog(BaseDialog):
    def __init__(self, emp: dict, progs: list, parent=None):
        super().__init__(parent, title="Редактирование сотрудника", min_width=520, min_height=400)
        self.emp = emp
        self.progs = progs

        bl = self.body_layout()
        form = QFormLayout(); form.setSpacing(8)

        from utils.field_validators import validate_required, validate_snils, validate_name

        self.last_name = ValidatedLineEdit(emp['last_name'],
            validator=lambda t: validate_required(t, "Фамилия") or validate_name(t))
        self.register_field(self.last_name)
        form.addRow("Фамилия:", self.last_name)

        self.first_name = ValidatedLineEdit(emp['first_name'],
            validator=lambda t: validate_required(t, "Имя") or validate_name(t))
        self.register_field(self.first_name)
        form.addRow("Имя:", self.first_name)

        self.middle_name = ValidatedLineEdit(emp['middle_name'], validator=validate_name)
        form.addRow("Отчество:", self.middle_name)

        self.snils = ValidatedLineEdit(emp['snils'], validator=validate_snils)
        form.addRow("СНИЛС:", self.snils)

        self.position = ValidatedLineEdit(emp['position'],
            validator=lambda t: validate_required(t, "Должность"))
        self.register_field(self.position)
        form.addRow("Должность:", self.position)

        prog_ids = [str(p['program_id']) for p in progs if p.get('need_training') == 1]
        self.programs = QLineEdit(",".join(prog_ids))
        form.addRow("Программы:", self.programs)

        bl.addLayout(form)
        btn_row = QHBoxLayout()
        save_btn = QPushButton("Сохранить"); save_btn.setObjectName("dialogPrimaryBtn")
        save_btn.clicked.connect(self.accept)
        cancel_btn = QPushButton("Отмена"); cancel_btn.setObjectName("dialogDangerBtn")
        cancel_btn.clicked.connect(self.reject)
        btn_row.addWidget(save_btn); btn_row.addStretch(); btn_row.addWidget(cancel_btn)
        bl.addLayout(btn_row)


# ──────── EmployeeSummaryTab ────────

class EmployeeSummaryTab(QWidget):
    _programs_cache = None

    @classmethod
    def _load_saved_settings(cls, data_dir):
        path = os.path.join(data_dir, "summary_programs.json")
        result = {'programs': None, 'b_period_3years': True}
        try:
            if os.path.exists(path):
                with open(path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                result['programs'] = [p for p in data.get('programs', []) if p in VALID_PROGRAMS]
                result['b_period_3years'] = data.get('b_period_3years', True)
        except (OSError, json.JSONDecodeError) as e:
            logger.debug("Failed to load summary_programs.json: %s", e)
        return result

    @classmethod
    def _save_settings(cls, data_dir, programs, b_period_3years=True):
        path = os.path.join(data_dir, "summary_programs.json")
        try:
            with open(path, 'w', encoding='utf-8') as f:
                json.dump({'programs': programs, 'b_period_3years': b_period_3years}, f)
        except OSError as e:
            logger.debug("Failed to save summary_programs.json: %s", e)

    def __init__(self, parent=None):
        super().__init__(parent)
        from utils.app_paths import get_app_data_dir
        self.data_dir = get_app_data_dir()
        saved = self._load_saved_settings(self.data_dir)
        self._selected_programs = saved['programs'] if saved['programs'] else DEFAULT_PROGRAMS.copy()
        self._b_period_3years = saved['b_period_3years']
        self._current_filter_status = "all"
        self._current_filter_program = "all"
        self._current_filter_position = ""
        self._problem_only = False

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(8, 8, 8, 8)
        main_layout.setSpacing(8)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)

        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)
        scroll_layout.setContentsMargins(0, 0, 0, 0)
        scroll_layout.setSpacing(8)

        def _make_hscroll(widget: QWidget) -> QScrollArea:
            sa = QScrollArea()
            sa.setWidgetResizable(False)
            sa.setFrameShape(QFrame.Shape.NoFrame)
            sa.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
            sa.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
            sa.setWidget(widget)
            h = widget.sizeHint().height()
            if h > 0:
                sa.setFixedHeight(h + 4)
            return sa

        scroll_layout.addWidget(self._build_stats())
        scroll_layout.addWidget(_make_hscroll(self._build_period_row()))
        scroll_layout.addWidget(_make_hscroll(self._build_report_row()))
        scroll_layout.addWidget(_make_hscroll(self._build_toolbar()))
        scroll_layout.addWidget(_make_hscroll(self._build_filters()))
        scroll_layout.addStretch()

        scroll.setWidget(scroll_widget)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)

        main_layout.addWidget(scroll)
        main_layout.addWidget(self._build_table_widget(), 1)

        self.refresh_table()

    # ── Stats Cards ────────────────────────────────────────

    def _build_stats(self):
        container = QWidget()
        container.setObjectName("statsContainer")
        layout = QHBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(10)

        self.stat_total = QLabel("0"); self.stat_total.setObjectName("statValue")
        self.stat_trained = QLabel("0"); self.stat_trained.setObjectName("statValueGreen")
        self.stat_untrained = QLabel("0"); self.stat_untrained.setObjectName("statValueRed")
        self.stat_expired = QLabel("0"); self.stat_expired.setObjectName("statValueYellow")
        self.stat_sync = QLabel("нет"); self.stat_sync.setObjectName("statValueInfo")

        for value_lbl, label, obj_name in [
            (self.stat_total, "Всего сотрудников", "statCardBlue"),
            (self.stat_trained, "Обучено сотрудников", "statCardGreen"),
            (self.stat_untrained, "Не обучено сотрудников", "statCardRed"),
            (self.stat_expired, "Просрочено сотрудников", "statCardYellow"),
            (self.stat_sync, "Актуальность данных", "statCardInfo"),
        ]:
            card = QFrame(); card.setObjectName(obj_name)
            cl = QVBoxLayout(card); cl.setSpacing(2); cl.setContentsMargins(8, 4, 8, 4)
            value_lbl.setObjectName("statValue")
            label_lbl = QLabel(label); label_lbl.setObjectName("statLabel")
            cl.addWidget(value_lbl, 0, Qt.AlignmentFlag.AlignCenter)
            cl.addWidget(label_lbl, 0, Qt.AlignmentFlag.AlignCenter)
            layout.addWidget(card)

        return container

    # ── Toolbar ────────────────────────────────────────────

    def _build_toolbar(self):
        w = QWidget(); w.setObjectName("toolbarContainer")
        layout = QHBoxLayout(w); layout.setContentsMargins(0, 0, 0, 0); layout.setSpacing(6)

        add_btn = QPushButton("+ Добавить сотрудника"); add_btn.setObjectName("toolbarPrimaryBtn")
        add_btn.clicked.connect(self._add_employee_dialog)
        layout.addWidget(add_btn)

        import_btn = QPushButton("Импорт из xlsx"); import_btn.setObjectName("toolbarBtn")
        import_btn.clicked.connect(self._import_xlsx)
        layout.addWidget(import_btn)

        export_btn = QPushButton("Экспорт в xlsx"); export_btn.setObjectName("toolbarBtn")
        export_btn.clicked.connect(lambda: self._export_xlsx(filtered=True))
        layout.addWidget(export_btn)

        export_all_btn = QPushButton("Экспорт (все)"); export_all_btn.setObjectName("toolbarBtn")
        export_all_btn.clicked.connect(lambda: self._export_xlsx(filtered=False))
        layout.addWidget(export_all_btn)

        self.query_btn = QPushButton("Запросить из реестра"); self.query_btn.setObjectName("toolbarSuccessBtn")
        self.query_btn.clicked.connect(self._query_reestr)
        layout.addWidget(self.query_btn)

        prog_btn = QPushButton("Выбрать программы"); prog_btn.setObjectName("toolbarPurpleBtn")
        prog_btn.clicked.connect(self._show_program_selector)
        layout.addWidget(prog_btn)

        layout.addStretch()

        delete_btn = QPushButton("Удалить данные"); delete_btn.setObjectName("toolbarDangerBtn")
        delete_btn.clicked.connect(self._delete_all_data)
        layout.addWidget(delete_btn)

        return w

    # ── Period settings row ───────────────────────────────

    def _build_period_row(self):
        w = QWidget(); w.setObjectName("periodRowContainer")
        layout = QHBoxLayout(w); layout.setContentsMargins(0, 0, 0, 0); layout.setSpacing(8)

        self.period_cb = QCheckBox("Обучение по программам В (№6-29) — 1 раз в 3 года")
        self.period_cb.setObjectName("periodCheckBox")
        self.period_cb.setChecked(self._b_period_3years)
        self.period_cb.toggled.connect(self._on_period_toggled)
        layout.addWidget(self.period_cb)
        layout.addStretch()

        return w

    def _on_period_toggled(self, checked):
        self._b_period_3years = checked
        self._save_settings(self.data_dir, self._selected_programs, self._b_period_3years)
        self.refresh_table()

    # ── Report row (plan / snapshot / trained) ────────────

    def _build_report_row(self):
        w = QWidget(); w.setObjectName("reportRowContainer")
        layout = QHBoxLayout(w); layout.setContentsMargins(0, 0, 0, 0); layout.setSpacing(8)

        plan_current_btn = QPushButton("План на текущий год"); plan_current_btn.setObjectName("planBtn")
        plan_current_btn.clicked.connect(lambda: self._generate_plan(current_year=True))
        layout.addWidget(plan_current_btn)

        plan_next_btn = QPushButton("План на след. год"); plan_next_btn.setObjectName("planBtn")
        plan_next_btn.clicked.connect(lambda: self._generate_plan(current_year=False))
        layout.addWidget(plan_next_btn)

        snapshot_btn = QPushButton("Текущая ситуация"); snapshot_btn.setObjectName("planBtnInfo")
        snapshot_btn.clicked.connect(self._show_current_snapshot)
        layout.addWidget(snapshot_btn)

        report_btn = QPushButton("Отчет по обученным"); report_btn.setObjectName("planBtnSuccess")
        report_btn.clicked.connect(self._generate_trained_report)
        layout.addWidget(report_btn)

        layout.addStretch()
        return w

    # ── Filters ───────────────────────────────────────────

    def _build_filters(self):
        w = QWidget(); w.setObjectName("filterContainer")
        layout = QHBoxLayout(w); layout.setContentsMargins(8, 4, 8, 4); layout.setSpacing(10)

        layout.addWidget(QLabel("Программа:"))
        self.filter_program = QComboBox(); self.filter_program.setObjectName("filterCombo")
        self.filter_program.addItem("Все", "all")
        for p in VALID_PROGRAMS:
            self.filter_program.addItem(f"{p}", p)
        self.filter_program.currentIndexChanged.connect(self._apply_filters)
        layout.addWidget(self.filter_program)

        layout.addWidget(QLabel("Статус:"))
        self.filter_status = QComboBox(); self.filter_status.setObjectName("filterCombo")
        self.filter_status.addItems(["Все", "Обучен", "Не обучен", "Просрочено"])
        self.filter_status.currentIndexChanged.connect(self._apply_filters)
        layout.addWidget(self.filter_status)

        layout.addWidget(QLabel("Должность:"))
        self.filter_position = QLineEdit(); self.filter_position.setObjectName("filterInput")
        self.filter_position.setPlaceholderText("Фильтр по должности")
        self.filter_position.textChanged.connect(self._apply_filters)
        layout.addWidget(self.filter_position)

        self.problem_cb = QCheckBox("Только проблемные"); self.problem_cb.setObjectName("filterCheck")
        self.problem_cb.toggled.connect(self._apply_filters)
        layout.addWidget(self.problem_cb)

        return w

    # ── Table ─────────────────────────────────────────────

    def _build_table_widget(self):
        self.table = QTableWidget()
        self.table.setObjectName("summaryTable")
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table.verticalHeader().setDefaultSectionSize(28)
        self.table.verticalHeader().setVisible(False)
        self.table.setSortingEnabled(True)

        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._show_context_menu)
        self.table.itemDoubleClicked.connect(self._on_item_double_click)

        return self.table

    def refresh_table(self):
        employees = EmployeesRepo.get_all()
        self._build_table(employees)
        self._update_stats()

    def _build_table(self, employees: List[dict]):
        self.table.setSortingEnabled(False)
        self.table.clear()
        self.table.setRowCount(0)

        programs = self._selected_programs[:6]
        num_prog_cols = len(programs) * SUB_COLUMNS
        num_cols = BASE_COLUMNS + num_prog_cols + 1

        self.table.setColumnCount(num_cols)

        pal = self.table.palette()
        hdr_bg = pal.color(QPalette.ColorRole.Highlight)
        hdr_fg = QColor("white")

        self.table.insertRow(0)
        base_headers = ["ФИО", "СНИЛС", "Должность"]
        header_font = QFont()
        header_font.setBold(True)
        for c in range(BASE_COLUMNS):
            item = QTableWidgetItem(base_headers[c])
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            item.setFlags(Qt.ItemFlag.ItemIsEnabled)
            item.setBackground(hdr_bg)
            item.setForeground(hdr_fg)
            item.setFont(header_font)
            self.table.setItem(0, c, item)

        for pi, p in enumerate(programs):
            col = BASE_COLUMNS + pi * SUB_COLUMNS
            item = QTableWidgetItem(f"№{p}: {PROGRAM_TITLES.get(p, '')}")
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            item.setFlags(Qt.ItemFlag.ItemIsEnabled)
            item.setBackground(hdr_bg)
            item.setForeground(hdr_fg)
            item.setFont(header_font)
            self.table.setItem(0, col, item)
            self.table.setSpan(0, col, 1, SUB_COLUMNS)

        item = QTableWidgetItem("")
        item.setFlags(Qt.ItemFlag.ItemIsEnabled)
        self.table.setItem(0, num_cols - 1, item)
        self.table.setColumnHidden(num_cols - 1, True)

        self.table.insertRow(1)
        for c in range(BASE_COLUMNS):
            item = QTableWidgetItem("")
            item.setFlags(Qt.ItemFlag.ItemIsEnabled)
            self.table.setItem(1, c, item)

        sub_labels = ["Потребность", "Дата", "Протокол", "Рег.№"]
        sub_font = QFont()
        sub_font.setBold(True)
        for pi, p in enumerate(programs):
            col = BASE_COLUMNS + pi * SUB_COLUMNS
            for j, label in enumerate(sub_labels):
                item = QTableWidgetItem(label)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                item.setFlags(Qt.ItemFlag.ItemIsEnabled)
                item.setBackground(hdr_bg)
                item.setForeground(hdr_fg)
                item.setFont(sub_font)
                self.table.setItem(1, col + j, item)

        self.table.setRowHeight(0, 36)
        self.table.setRowHeight(1, 30)

        status_filter = self.filter_status.currentText()
        prog_filter = self.filter_program.currentData()
        pos_filter = self.filter_position.text().strip().lower()
        problem_only = self.problem_cb.isChecked()

        for emp in employees:
            if pos_filter:
                emp_pos = emp['position'].lower()
                if pos_filter not in emp_pos:
                    continue

            progs = EmployeeProgramsRepo.get_by_employee(emp['id'])
            if prog_filter != "all":
                has_prog = any(str(p['program_id']) == prog_filter and p.get('need_training') == 1 for p in progs)
                if not has_prog:
                    continue
            prog_map = {str(p['program_id']): p for p in progs}

            row_programs = {}
            overall_status = None
            for p in programs:
                pd = prog_map.get(p)
                if not pd or pd.get('need_training') != 1:
                    row_programs[p] = {'need_training': 0, 'exam_date': '', 'protocol': '', 'base_no': '', 'status': ''}
                    continue
                stored_status = pd.get('status', 'not_trained')
                prog_id = int(p)
                effective_status = get_dynamic_status(
                    stored_status, pd.get('exam_date', ''),
                    prog_id, self._b_period_3years
                )
                row_programs[p] = { 'need_training': pd.get('need_training', 0),
                    'exam_date': pd.get('exam_date', ''), 'protocol': pd.get('protocol', ''),
                    'base_no': pd.get('base_no', ''), 'status': effective_status,}
                if effective_status == 'not_trained':
                    overall_status = 'not_trained'
                elif effective_status == 'expired' and overall_status != 'not_trained':
                    overall_status = 'expired'
                elif effective_status == 'trained' and overall_status is None:
                    overall_status = 'trained'

            if status_filter != "Все":
                status_map = {"Обучен": "trained", "Не обучен": "not_trained", "Просрочено": "expired"}
                mapped = status_map.get(status_filter, "")
                if overall_status != mapped:
                    continue

            if problem_only and overall_status in (None, 'trained'):
                continue

            row_idx = self.table.rowCount()
            self.table.insertRow(row_idx)

            fio = f"{emp['last_name']} {emp['first_name']} {emp['middle_name']}".strip()
            row_data = [fio, emp['snils'], emp['position']]

            for p in programs:
                pd = row_programs.get(p, {})
                need = pd.get('need_training', 0)
                need_text = "Да" if need == 1 else "Нет"
                row_data.extend([need_text, pd.get('exam_date', ''), pd.get('protocol', ''), pd.get('base_no', '')])

            col_idx = 0
            for val in row_data:
                text = str(val) if val is not None else ""
                item = QTableWidgetItem(text)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.table.setItem(row_idx, col_idx, item)
                col_idx += 1

            emp_id_item = QTableWidgetItem(str(emp['id']))
            self.table.setItem(row_idx, col_idx, emp_id_item)
            self.table.item(row_idx, col_idx).setData(Qt.ItemDataRole.UserRole, emp['id'])

            for pi, p in enumerate(programs):
                pd = row_programs.get(p, {})
                need = pd.get('need_training', 0)
                if need != 1:
                    continue
                s = pd.get('status', 'not_trained')
                col_start = BASE_COLUMNS + pi * SUB_COLUMNS
                if s == 'trained':
                    color = QColor("#d4edda")
                elif s == 'expired':
                    color = QColor("#fff3cd")
                else:
                    color = QColor("#f8d7da")
                for sub in range(SUB_COLUMNS):
                    ci = col_start + sub
                    it = self.table.item(row_idx, ci)
                    if it:
                        it.setBackground(color)

        self.table.setSortingEnabled(True)

    def _get_employee_status(self, emp_id: int, programs) -> str:
        has_trained = False
        has_expired = False
        for p in programs:
            if p.get('need_training') != 1:
                continue
            stored_status = p.get('status', 'not_trained')
            s = get_dynamic_status(
                stored_status, p.get('exam_date', ''),
                p.get('program_id', 0), self._b_period_3years
            )
            if s == 'not_trained':
                return 'not_trained'
            if s == 'expired':
                has_expired = True
            elif s == 'trained':
                has_trained = True
        if has_expired:
            return 'expired'
        if has_trained:
            return 'trained'
        return 'not_trained'

    def _update_stats(self):
        trained = not_trained = expired = 0
        emp_count = 0
        for emp in EmployeesRepo.get_all():
            progs = EmployeeProgramsRepo.get_by_employee(emp['id'])
            need_progs = [p for p in progs if p.get('need_training') == 1]
            if not need_progs:
                continue
            emp_count += 1
            status = self._get_employee_status(emp['id'], progs)
            if status == 'trained':
                trained += 1
            elif status == 'expired':
                expired += 1
            else:
                not_trained += 1
        self.stat_total.setText(str(emp_count))
        self.stat_trained.setText(str(trained))
        self.stat_untrained.setText(str(not_trained))
        self.stat_expired.setText(str(expired))

        last_sync = ""
        rows = DatabaseManager.get_instance().fetchone(
            "SELECT MAX(last_sync) as ls FROM employees WHERE last_sync IS NOT NULL"
        )
        if rows and rows['ls']:
            parts = rows['ls'].split(' ', 1)
            last_sync = '\n'.join(parts) if len(parts) > 1 else rows['ls']
        self.stat_sync.setText(last_sync or "нет")

    def _apply_filters(self):
        self.refresh_table()

    # ── Context menu ───────────────────────────────────────

    def _show_context_menu(self, position):
        idx = self.table.indexAt(position)
        if not idx.isValid():
            return
        row = idx.row()
        if row < 2:
            return

        menu = QMenu(self)
        menu.setObjectName("summaryCtxMenu")

        edit_action = menu.addAction("Редактировать")
        query_action = menu.addAction("Запросить из реестра")
        delete_action = menu.addAction("Удалить")

        action = menu.exec(self.table.mapToGlobal(position))
        hidden_col = self.table.columnCount() - 1
        emp_id_item = self.table.item(row, hidden_col)
        emp_id = emp_id_item.data(Qt.ItemDataRole.UserRole) if emp_id_item else None

        if action == edit_action:
            if emp_id:
                self._edit_employee(emp_id)
        elif action == query_action:
            if emp_id:
                self._query_single(emp_id)
        elif action == delete_action:
            if emp_id:
                reply = QMessageBox.question(
                    self, "Подтверждение", "Удалить сотрудника из сводки?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                )
                if reply == QMessageBox.StandardButton.Yes:
                    EmployeesRepo.delete(emp_id)
                    self.refresh_table()

    def _on_item_double_click(self, item):
        if not item:
            return
        row = item.row()
        if row < 2:
            return
        col = item.column()
        hidden_idx = self.table.columnCount() - 1

        programs = self._selected_programs[:6]
        for pi, p in enumerate(programs):
            col_start = BASE_COLUMNS + pi * SUB_COLUMNS
            if col == col_start:
                emp_id = self.table.item(row, hidden_idx).data(Qt.ItemDataRole.UserRole) if self.table.item(row, hidden_idx) else None
                if not emp_id:
                    return
                current = self.table.item(row, col).text() if self.table.item(row, col) else "Нет"
                new_val = "Да" if current != "Да" else "Нет"
                need = 1 if new_val == "Да" else 0
                EmployeeProgramsRepo.update_need_training(emp_id, int(p), need)
                self.refresh_table()
                return

    # ── Dialogs ────────────────────────────────────────────

    def _add_employee_dialog(self):
        dialog = EmployeeAddDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            last_name = dialog.last_name.text().strip()
            first_name = dialog.first_name.text().strip()
            middle_name = dialog.middle_name.text().strip()
            snils = dialog.snils.text().strip()
            position = dialog.position.text().strip()
            programs_str = dialog.programs.text().strip()

            if not snils:
                QMessageBox.warning(self, "Ошибка", "Заполните СНИЛС")
                return

            snils_clean = snils.replace('-', '').replace(' ', '')
            if not snils_clean.isdigit() or len(snils_clean) != 11:
                QMessageBox.warning(self, "Ошибка", "СНИЛС должен содержать 11 цифр")
                return

            snils_fmt = f"{snils_clean[:3]}-{snils_clean[3:6]}-{snils_clean[6:9]} {snils_clean[9:]}"
            programs = [p.strip() for p in programs_str.split(',') if p.strip() and p.strip() in VALID_PROGRAMS]

            emp_data = { 'snils': snils_fmt, 'last_name': last_name, 'first_name': first_name,
                'middle_name': middle_name, 'position': position,
                'required_programs': ';'.join(programs),}
            emp_id = EmployeesRepo.upsert(emp_data)

            for p in programs:
                EmployeeProgramsRepo.upsert(emp_id, {'program_id': int(p), 'need_training': 1})

            self.refresh_table()
            QMessageBox.information(self, "Успех", "Сотрудник добавлен")

    def _edit_employee(self, emp_id):
        emp = EmployeesRepo.get_by_id(emp_id)
        if not emp:
            return
        progs = EmployeeProgramsRepo.get_by_employee(emp_id)
        old_progs = {str(p['program_id']): p for p in progs}
        dialog = EmployeeEditDialog(emp, progs, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            new_progs = [p.strip() for p in dialog.programs.text().split(',') if p.strip() in VALID_PROGRAMS]
            EmployeesRepo.upsert({ 'snils': dialog.snils.text().strip(),
                'last_name': dialog.last_name.text().strip(), 'first_name': dialog.first_name.text().strip(),
                'middle_name': dialog.middle_name.text().strip(), 'position': dialog.position.text().strip(),
                'required_programs': ';'.join(new_progs),})
            EmployeeProgramsRepo.delete_by_employee(emp_id)
            for p in new_progs:
                old = old_progs.get(p, {})
                EmployeeProgramsRepo.upsert(emp_id, {
                    'program_id': int(p),
                    'need_training': 1,
                    'exam_date': old.get('exam_date'),
                    'protocol': old.get('protocol'),
                    'base_no': old.get('base_no'),
                    'result': old.get('result'),
                })
            self.refresh_table()

    def _delete_all_data(self):
        reply = QMessageBox.question(
            self, "Подтверждение",
            "Удалить ВСЕ данные из сводки по сотрудникам?\nЭто действие необратимо.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            EmployeesRepo.clear()
            self.refresh_table()
            QMessageBox.information(self, "Успех", "Все данные удалены")

    # ── Program selector ───────────────────────────────────

    def _show_program_selector(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Выбор программ для отображения")
        dialog.setMinimumSize(500, 500)

        lw = QListWidget()
        selected = set(self._selected_programs)
        for p in VALID_PROGRAMS:
            item = QListWidgetItem(f"{p}: {PROGRAM_TITLES.get(p, '')}")
            item.setData(Qt.ItemDataRole.UserRole, p)
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            item.setCheckState(Qt.CheckState.Checked if p in selected else Qt.CheckState.Unchecked)
            lw.addItem(item)

        layout = QVBoxLayout(dialog)
        layout.addWidget(QLabel("Отметьте программы для отображения (макс. 6):"))
        layout.addWidget(lw)

        btn_row = QHBoxLayout()
        ok_btn = QPushButton("Применить"); ok_btn.setObjectName("dialogPrimaryBtn")
        def apply():
            checked = []
            for i in range(lw.count()):
                item = lw.item(i)
                if item.checkState() == Qt.CheckState.Checked:
                    checked.append(item.data(Qt.ItemDataRole.UserRole))
            if not checked:
                QMessageBox.warning(self, "Ошибка", "Выберите хотя бы одну программу")
                return
            if len(checked) > 6:
                QMessageBox.warning(self, "Ошибка", "Максимум 6 программ одновременно")
                return
            self._selected_programs = checked
            self._save_settings(self.data_dir, checked, self._b_period_3years)
            dialog.accept()
            self.refresh_table()
        ok_btn.clicked.connect(apply)
        cancel_btn = QPushButton("Отмена"); cancel_btn.setObjectName("dialogDangerBtn")
        cancel_btn.clicked.connect(dialog.reject)
        btn_row.addWidget(ok_btn); btn_row.addStretch(); btn_row.addWidget(cancel_btn)
        layout.addLayout(btn_row)
        dialog.exec()

    # ── Import / Export ────────────────────────────────────

    def _import_xlsx(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Выберите XLSX файл", "", "Excel Files (*.xlsx)")
        if not file_path:
            return
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active

            headers = [str(ws.cell(row=1, column=c).value or '').strip() for c in range(1, ws.max_column + 1)]
            col_map = {h: i for i, h in enumerate(headers)}

            if 'СНИЛС' not in col_map:
                QMessageBox.warning(self, "Ошибка", "Не найден столбец 'СНИЛС'")
                return

            snils_col = col_map['СНИЛС']
            last_name_col = col_map.get('Фамилия')
            first_name_col = col_map.get('Имя')
            middle_name_col = col_map.get('Отчество')
            position_col = col_map.get('Должность')
            fio_col = col_map.get('ФИО')

            program_cols = {}
            for p in VALID_PROGRAMS:
                need_key = f"программа #{p} Потребность"
                date_key = f"программа #{p} Дата обучения"
                proto_key = f"программа #{p} Протокол"
                reg_key = f"программа #{p} Рег.№"
                if need_key in col_map:
                    program_cols[p] = { 'need': col_map[need_key],
                        'date': col_map.get(date_key), 'protocol': col_map.get(proto_key),
                        'reg': col_map.get(reg_key),}

            hidden_col = col_map.get('Потребность в обучении по программам')

            if not program_cols and hidden_col is None:
                QMessageBox.warning(self, "Ошибка", "Не найдены столбцы с потребностью в обучении по программам.\n\nОжидаются заголовки вида: 'программа #N Потребность' или 'Потребность в обучении по программам'")
                return

            imported = 0
            for row_num in range(2, ws.max_row + 1):
                snils_raw = str(ws.cell(row=row_num, column=snils_col + 1).value or '').strip()
                if not snils_raw:
                    continue
                snils_clean = snils_raw.replace('-', '').replace(' ', '')
                if not snils_clean.isdigit() or len(snils_clean) != 11:
                    logger.info("Import: row %d skipped — invalid SNILS", row_num)
                    continue
                snils_fmt = f"{snils_clean[:3]}-{snils_clean[3:6]}-{snils_clean[6:9]} {snils_clean[9:]}"

                last_name = str(ws.cell(row=row_num, column=(last_name_col if last_name_col is not None else 0) + 1).value or '').strip() if last_name_col is not None else ''
                first_name = str(ws.cell(row=row_num, column=(first_name_col if first_name_col is not None else 0) + 1).value or '').strip() if first_name_col is not None else ''
                middle_name = str(ws.cell(row=row_num, column=(middle_name_col if middle_name_col is not None else 0) + 1).value or '').strip() if middle_name_col is not None else ''
                position = str(ws.cell(row=row_num, column=(position_col if position_col is not None else 0) + 1).value or '').strip() if position_col is not None else ''
                if (not last_name or not first_name) and fio_col is not None:
                    fio_raw = str(ws.cell(row=row_num, column=fio_col + 1).value or '').strip()
                    if fio_raw:
                        parts = fio_raw.split()
                        if len(parts) >= 1: last_name = parts[0]
                        if len(parts) >= 2: first_name = parts[1]
                        if len(parts) >= 3: middle_name = ' '.join(parts[2:])

                required_programs = set()
                if hidden_col is not None:
                    hidden_val = str(ws.cell(row=row_num, column=hidden_col + 1).value or '').strip()
                    if hidden_val:
                        parts = [p.strip() for p in hidden_val.replace(';', ',').split(',')]
                        for p in parts:
                            if p in VALID_PROGRAMS:
                                required_programs.add(p)

                for p, cols in program_cols.items():
                    val = str(ws.cell(row=row_num, column=cols['need'] + 1).value or '').strip().lower()
                    if val in ('1', 'да', 'true', 'yes'):
                        required_programs.add(p)

                emp_id = EmployeesRepo.upsert({ 'snils': snils_fmt,
                    'last_name': last_name, 'first_name': first_name, 'middle_name': middle_name,
                    'position': position,
                    'required_programs': ';'.join(sorted(required_programs)),})
                for p in required_programs:
                    prog_data = {'program_id': int(p), 'need_training': 1}
                    if p in program_cols:
                        cols = program_cols[p]
                        if cols['date'] is not None:
                            date_val = ws.cell(row=row_num, column=cols['date'] + 1).value
                            if date_val is not None:
                                from datetime import datetime as _dt
                                if isinstance(date_val, _dt):
                                    prog_data['exam_date'] = date_val.strftime('%d.%m.%Y')
                                else:
                                    prog_data['exam_date'] = str(date_val).strip()
                        if cols['protocol'] is not None:
                            pval = ws.cell(row=row_num, column=cols['protocol'] + 1).value
                            if pval is not None: prog_data['protocol'] = str(pval).strip()
                        if cols['reg'] is not None:
                            rval = ws.cell(row=row_num, column=cols['reg'] + 1).value
                            if rval is not None: prog_data['base_no'] = str(rval).strip()
                    EmployeeProgramsRepo.upsert(emp_id, prog_data)
                imported += 1

            if imported == 0:
                msg = "Не найдено сотрудников для импорта."
                safe_message_box(self, QMessageBox.Icon.Information, "Результат", msg)
            else:
                safe_message_box(self, QMessageBox.Icon.Information, "Успех", f"Импортировано: {imported} сотрудников")
            self.refresh_table()
        except Exception as e:
            logger.exception("Ошибка импорта XLSX")
            safe_message_box(self, QMessageBox.Icon.Warning, "Ошибка", f"Ошибка импорта: {e}")

    def _export_xlsx(self, filtered=True):
        reply = QMessageBox.warning(
            self, "Экспорт данных",
            "⚠️ Файл будет содержать персональные данные (ФИО, СНИЛС).\n"
            "Убедитесь, что файл сохраняется на зашифрованном диске (BitLocker).",
            QMessageBox.StandardButton.Ok | QMessageBox.StandardButton.Cancel
        )
        if reply != QMessageBox.StandardButton.Ok:
            return
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить XLSX", "Сводка_сотрудников.xlsx", "Excel Files (*.xlsx)")
        if not file_path:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            employees = EmployeesRepo.get_all()
            log_audit("EXPORT_XLSX", f"employees={len(employees)}")
            wb = Workbook()
            ws = wb.active
            ws.title = "Сводка"

            programs = self._selected_programs[:6] if filtered else VALID_PROGRAMS

            headers = ["ФИО", "СНИЛС", "Должность"]
            for p in programs:
                headers.append(f"программа #{p} Потребность")
                headers.append(f"программа #{p} Дата обучения")
                headers.append(f"программа #{p} Протокол")
                headers.append(f"программа #{p} Рег.№")
            ws.append(headers)

            hf = Font(bold=True, color="FFFFFF")
            hfill = PatternFill(start_color="4169E1", end_color="4169E1", fill_type="solid")
            for cell in ws[1]:
                cell.font = hf; cell.fill = hfill; cell.alignment = Alignment(horizontal="center")

            for emp in employees:
                progs = EmployeeProgramsRepo.get_by_employee(emp['id'])
                prog_map = {str(p['program_id']): p for p in progs}
                fio = f"{emp['last_name']} {emp['first_name']} {emp['middle_name']}".strip()
                row_data = [fio, emp['snils'], emp['position']]
                for p in programs:
                    pd = prog_map.get(p, {})
                    need = "Да" if pd.get('need_training') == 1 else "Нет"
                    row_data.extend([need, pd.get('exam_date', ''), pd.get('protocol', ''), pd.get('base_no', '')])
                ws.append(row_data)

            for col in ws.columns:
                max_len = 0
                for cell in col:
                    if cell.value: max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

            wb.save(file_path)
            safe_message_box(self, QMessageBox.Icon.Information, "Успех", f"Файл сохранён:\n{file_path}")
        except Exception as e:
            logger.exception("Export error")
            safe_message_box(self, QMessageBox.Icon.Warning, "Ошибка", f"Ошибка экспорта: {e}")

    # ── API ──────────────────────────────────────────────

    def _query_reestr(self):
        api_key = load_api_key(self.data_dir)
        if not api_key:
            QMessageBox.warning(self, "Ошибка", "API ключ не найден. Сохраните ключ на вкладке 'Передача данных'")
            return
        employees = EmployeesRepo.get_all()
        if not employees:
            QMessageBox.information(self, "Информация", "Нет сотрудников для запроса")
            return
        proxy_settings = load_proxy_settings(self.data_dir)

        self.query_btn.setEnabled(False)
        self.query_btn.setText("Запрос...")

        self._api_thread = ApiQueryThread(employees, api_key, proxy_settings)
        self._api_thread.finished.connect(self._on_query_finished)
        self._api_thread.error_signal.connect(lambda msg: logger.error("%s", filter_sensitive_text(str(msg))))
        self._api_thread.progress.connect(lambda c, t: self.query_btn.setText(f"Запрос... {c}/{t}"))
        self._api_thread.start()

    def _on_query_finished(self, updated, errors):
        self.query_btn.setEnabled(True)
        self.query_btn.setText("Запросить из реестра")
        if errors > 0:
            safe_message_box(self, QMessageBox.Icon.Warning, "Результат", f"Обновлено: {updated}\nОшибок: {errors}")
        else:
            safe_message_box(self, QMessageBox.Icon.Information, "Успех", f"Обновлено: {updated} сотрудников")
        self.refresh_table()

    def _query_single(self, emp_id):
        api_key = load_api_key(self.data_dir)
        if not api_key:
            QMessageBox.warning(self, "Ошибка", "API ключ не найден")
            return
        emp = EmployeesRepo.get_by_id(emp_id)
        if not emp:
            return
        proxy_settings = load_proxy_settings(self.data_dir)
        snils_clean = emp['snils'].replace('-', '').replace(' ', '')
        try:
            result = get_by_snils(api_key, snils_clean, proxy_settings=proxy_settings)
            if result.get("success"):
                records = result.get("records", [])
                best = {}
                for rec in records:
                    prog_id = rec.get('learnProgramId', '')
                    if not prog_id: continue
                    exam_date = _normalize_api_date(rec.get('Date', ''))
                    if prog_id not in best or (exam_date and _dmy_gt(exam_date, _normalize_api_date(best[prog_id].get('Date', '')))):
                        best[prog_id] = rec
                updated = 0
                for prog_id, rec in best.items():
                    is_passed = rec.get('isPassed', '')
                    result_val = 1 if is_passed and is_passed.lower() in ('true', '1', 'да') else 0
                    try:
                        EmployeeProgramsRepo.update_from_api(
                            emp_id, int(prog_id),
                            _normalize_api_date(rec.get('Date', '')), rec.get('ProtocolNumber', ''),
                            rec.get('baseNo', ''), result_val)
                        updated += 1
                    except (ValueError, TypeError): pass
                if records:
                    first_rec = records[0]
                    if not emp.get('last_name') or not emp.get('first_name') or not emp.get('position'):
                        updates = {}
                        if not emp.get('last_name'): updates['last_name'] = first_rec.get('LastName', '')
                        if not emp.get('first_name'): updates['first_name'] = first_rec.get('FirstName', '')
                        if not emp.get('middle_name'): updates['middle_name'] = first_rec.get('MiddleName', '')
                        if not emp.get('position'): updates['position'] = first_rec.get('Position', '')
                        if updates:
                            updates['snils'] = emp['snils']
                            EmployeesRepo.upsert(updates)
                EmployeesRepo.update_sync(emp_id, datetime.now().strftime('%d.%m.%Y %H:%M:%S'))
                safe_message_box(self, QMessageBox.Icon.Information, "Успех", f"Обновлено программ: {updated}")
                self.refresh_table()
            else:
                safe_message_box(self, QMessageBox.Icon.Warning, "Ошибка", result.get("error", "Неизвестная ошибка"))
        except Exception as e:
            logger.exception("Registry query error")
            safe_message_box(self, QMessageBox.Icon.Warning, "Ошибка", f"Ошибка запроса: {e}")

    # ── Reports ──────────────────────────────────────────

    def _get_program_data_for_status(self, emp: dict, programs: list, status: str, today: datetime, plan_year: int = None):
        """PERFORMANCE: найти релевантную программу для статуса сотрудника (employee-level)."""
        need_progs = [p for p in programs if p.get('need_training') == 1]
        if not need_progs:
            return None

        # Find the worst program matching employee status
        best_prog = None
        best_exam = ''
        best_expiry = ''
        best_prog_id = ''

        for p in need_progs:
            prog_id = p.get('program_id', 0)
            stored_status = p.get('status', 'not_trained')
            last_exam = p.get('exam_date', '')
            prog_status = get_dynamic_status(stored_status, last_exam, prog_id, self._b_period_3years)

            exp = ''
            if prog_status == 'trained' and last_exam:
                try:
                    ed = compute_expiry_date(last_exam, prog_id, self._b_period_3years)
                    exp = ed.strftime('%d.%m.%Y')
                except (ValueError, IndexError):
                    pass
            elif prog_status == 'not_trained':
                exp = (today + relativedelta(days=60)).strftime('%d.%m.%Y')
            elif prog_status == 'expired':
                exp = (today + relativedelta(days=30)).strftime('%d.%m.%Y')

            if prog_status == status:
                if best_prog is None:
                    best_prog = p
                    best_prog_id = str(prog_id)
                    best_exam = last_exam
                    best_expiry = exp
                elif best_expiry and exp and _dmy_gt(best_expiry, exp):
                    best_prog = p
                    best_prog_id = str(prog_id)
                    best_exam = last_exam
                    best_expiry = exp

        if best_prog is None and need_progs:
            best_prog = need_progs[0]
            best_prog_id = str(best_prog.get('program_id', ''))
            best_exam = best_prog.get('exam_date', '')

        return best_prog_id, best_exam, best_expiry

    def _show_current_snapshot(self):
        today = datetime.now()
        plan_data = []
        reason_map = {'not_trained': 'Не обучен', 'expired': 'Просрочено', 'trained': 'Обучен'}
        priority_map = {'not_trained': 'Высокий', 'expired': 'Высокий', 'trained': 'Низкий'}

        for emp in EmployeesRepo.get_all():
            progs = EmployeeProgramsRepo.get_by_employee(emp['id'])
            need_progs = [p for p in progs if p.get('need_training') == 1]
            if not need_progs:
                continue

            status = self._get_employee_status(emp['id'], progs)
            prog_data = self._get_program_data_for_status(emp, progs, status, today)

            if prog_data:
                prog_id, exam_date, expiry = prog_data
            else:
                prog_id = str(need_progs[0].get('program_id', ''))
                exam_date = need_progs[0].get('exam_date', '')
                expiry = ''

            plan_data.append({
                'last_name': emp['last_name'], 'first_name': emp['first_name'],
                'middle_name': emp['middle_name'], 'snils': emp['snils'],
                'position': emp['position'], 'program': prog_id,
                'last_exam_date': exam_date, 'expiry_date': expiry,
                'reason': reason_map.get(status, ''),
                'priority': priority_map.get(status, ''),
            })

        plan_data.sort(key=lambda x: (
            {"Высокий": 0, "Средний": 1, "Низкий": 2}.get(x['priority'], 3), x['last_name']))

        if not plan_data:
            QMessageBox.information(self, "Информация", "Нет данных для отображения")
            return

        log_audit("EXPORT_SNAPSHOT", f"employees={len(plan_data)}")

        title = f"Текущая ситуация на {today.strftime('%d.%m.%Y')}"
        plan_dlg = PlanDialog(plan_data, title, self)
        plan_dlg.exec()

    def _generate_trained_report(self):
        year = datetime.now().year

        categories = [
            ('А', [3]), ('Б', [4]),
            ('В', list(range(6, 30))),
            ('ОПП', [1]), ('СИЗ', [2]),
        ]

        db = DatabaseManager.get_instance()
        all_programs = db.fetchall(
            "SELECT program_id, exam_date FROM employee_programs WHERE need_training = 1"
        )

        report_data = {}
        for cat_name, prog_ids in categories:
            count_liable = 0
            count_trained = 0
            for row in all_programs:
                if row['program_id'] not in prog_ids: continue
                exam_date = row['exam_date']
                if not exam_date or not exam_date.strip():
                    count_liable += 1
                    continue
                try:
                    exam_date_str = exam_date.split()[0]
                    exam_year = datetime.strptime(exam_date_str, '%d.%m.%Y').year
                    if exam_year == year:
                        count_trained += 1
                        count_liable += 1
                    else:
                        period = get_training_period_years(row['program_id'], self._b_period_3years)
                        if exam_year == year - period:
                            count_liable += 1
                except (ValueError, IndexError):
                    count_liable += 1
                    continue
            report_data[cat_name] = (count_liable, count_trained)

        dialog = QDialog(self)
        dialog.setWindowTitle(f"Отчет по обученным за {year} год")
        dialog.setMinimumSize(700, 350)

        layout = QVBoxLayout(dialog)
        total_trained = sum(c[1] for c in report_data.values())
        log_audit("EXPORT_TRAINED_REPORT", f"employees={total_trained}")

        title = QLabel(f"Отчет по обученным за {year} год")
        title.setObjectName("reportTitleLabel")
        layout.addWidget(title)

        table = QTableWidget()
        table.setColumnCount(3)
        table.setHorizontalHeaderLabels(["Программа", "Подлежат обучению\nв текущем году", "Обучено\nв текущем году"])
        table.setAlternatingRowColors(True)
        table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        table.horizontalHeader().setDefaultSectionSize(200)
        table.verticalHeader().setVisible(False)

        for cat_name, (count_liable, count_trained) in report_data.items():
            row = table.rowCount()
            table.insertRow(row)
            for c, v in enumerate([cat_name, str(count_liable), str(count_trained)]):
                item = QTableWidgetItem(v)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                table.setItem(row, c, item)

        layout.addWidget(table)

        btn_layout = QHBoxLayout()
        close_btn = QPushButton("Закрыть"); close_btn.setObjectName("dialogDangerBtn")
        close_btn.clicked.connect(dialog.close)
        btn_layout.addStretch(); btn_layout.addWidget(close_btn)
        layout.addLayout(btn_layout)
        dialog.exec()

    def _generate_plan(self, current_year=True):
        year = datetime.now().year if current_year else datetime.now().year + 1
        plan_year = year
        today = datetime.now()

        dialog = QDialog(self)
        dialog.setWindowTitle(f"Параметры формирования плана на {plan_year} год")
        dialog.setMinimumSize(400, 250)

        layout = QVBoxLayout(dialog)
        layout.addWidget(QLabel(f"Параметры формирования плана на {plan_year} год:"))

        cb_not_trained = QCheckBox("Включать не обученных"); cb_not_trained.setChecked(True)
        layout.addWidget(cb_not_trained)
        cb_expired = QCheckBox("Включать просроченных"); cb_expired.setChecked(True)
        layout.addWidget(cb_expired)
        cb_expiring = QCheckBox("Включать истекающих в планируемом году"); cb_expiring.setChecked(True)
        layout.addWidget(cb_expiring)
        cb_failed = QCheckBox("Включать неуспешно прошедших"); cb_failed.setChecked(True)
        layout.addWidget(cb_failed)

        btn_layout = QHBoxLayout()
        generate_btn = QPushButton("Сформировать"); generate_btn.setObjectName("dialogPrimaryBtn")
        cancel_btn = QPushButton("Отмена"); cancel_btn.setObjectName("dialogDangerBtn")
        cancel_btn.clicked.connect(dialog.reject)

        plan_data = []

        def do_generate():
            nonlocal plan_data
            employees = EmployeesRepo.get_all()
            for emp in employees:
                progs = EmployeeProgramsRepo.get_by_employee(emp['id'])
                status = self._get_employee_status(emp['id'], progs)

                include = False
                reason = None
                priority = None
                expiry = ""
                prog_id = ""

                if status == 'not_trained':
                    if cb_not_trained.isChecked():
                        include = True
                        reason = "Не обучен"
                        priority = "Высокий"
                        expiry = (today + relativedelta(days=60)).strftime('%d.%m.%Y')
                elif status == 'expired':
                    if cb_expired.isChecked():
                        include = True
                        reason = "Просрочено"
                        priority = "Высокий"
                        expiry = (today + relativedelta(days=30)).strftime('%d.%m.%Y')
                else:
                    if cb_expiring.isChecked():
                        need_progs = [p for p in progs if p.get('need_training') == 1 and p.get('exam_date', '')]
                        for p in need_progs:
                            pid = p.get('program_id', 0)
                            last_exam = p.get('exam_date', '')
                            try:
                                expiry_date = compute_expiry_date(last_exam, pid, self._b_period_3years)
                                if expiry_date.year == plan_year:
                                    include = True
                                    reason = "Истекает срок действия"
                                    priority = "Средний"
                                    expiry = expiry_date.strftime('%d.%m.%Y')
                                    prog_id = str(pid)
                                    break
                            except (ValueError, IndexError):
                                pass

                if not include:
                    failed_progs = [p for p in progs if p.get('need_training') == 1 and p.get('result') == 0]
                    if failed_progs and cb_failed.isChecked():
                        include = True
                        reason = "Неудовлетворительный результат"
                        priority = "Высокий"
                        expiry = (today + relativedelta(days=60)).strftime('%d.%m.%Y')
                        fp = failed_progs[0]
                        if not prog_id:
                            prog_id = str(fp.get('program_id', ''))

                if include:
                    prog_data = self._get_program_data_for_status(emp, progs, status, today, plan_year)
                    if prog_data:
                        prog_id_from_data, last_exam, _ = prog_data
                        if not prog_id:
                            prog_id = prog_id_from_data
                    else:
                        last_exam = ''
                        if not prog_id:
                            need_progs = [p for p in progs if p.get('need_training') == 1]
                            prog_id = str(need_progs[0].get('program_id', '')) if need_progs else ''
                            last_exam = need_progs[0].get('exam_date', '') if need_progs else ''

                    plan_data.append({
                        'last_name': emp['last_name'], 'first_name': emp['first_name'],
                        'middle_name': emp['middle_name'], 'snils': emp['snils'],
                        'position': emp['position'], 'program': prog_id,
                        'last_exam_date': last_exam if prog_id else '',
                        'expiry_date': expiry,
                        'reason': reason, 'priority': priority,
                    })

            n = sum(1 for x in plan_data if x['reason'] == 'Не обучен')
            e = sum(1 for x in plan_data if x['reason'] == 'Просрочено')
            t = sum(1 for x in plan_data if x['reason'] == 'Истекает срок действия')
            log_audit("EXPORT_PLAN", f"employees={len(plan_data)}, not_trained={n}, expired={e}, trained={t}")

        plan_data.sort(key=lambda x: (
            {"Высокий": 0, "Средний": 1, "Низкий": 2}.get(x['priority'], 3), x['last_name']))
        dialog.accept()

        generate_btn.clicked.connect(do_generate)
        btn_layout.addWidget(generate_btn); btn_layout.addStretch(); btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)

        if dialog.exec() == QDialog.DialogCode.Rejected or not plan_data:
            if not plan_data:
                QMessageBox.information(self, "Информация", "Нет сотрудников для включения в план")
            return

        title = f"План обучения на {plan_year} год"
        plan_dlg = PlanDialog(plan_data, title, self)
        plan_dlg.exec()
