from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLineEdit, QPushButton,
    QLabel, QMessageBox, QFileDialog, QTableView, QHeaderView,
    QAbstractItemView, QMenu, QFormLayout, QComboBox
)
from utils.dialog_base import BaseDialog
from PySide6.QtCore import Qt, QSortFilterProxyModel
from PySide6.QtWidgets import QDialog
from PySide6.QtGui import QColor
from datetime import datetime
import os
import logging

from exporters.xml_exporter import export_to_xml
from db.workers_data_repo import WorkersDataRepo
from utils.crypto import decrypt_data, hash_for_search
from utils.table_models import DataViewTableModel, MultiColumnFilterProxyModel, FIELD_KEYS, COLUMN_LABELS

logger = logging.getLogger(__name__)


class DataViewTab(QWidget):
    def __init__(self):
        super().__init__()
        self._setup_ui()
        self._connect_signals()
        self._load_all_data()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(8)

        toolbar = QHBoxLayout()
        toolbar.setSpacing(6)

        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("\U0001F50D Поиск...")
        self.search_edit.setClearButtonEnabled(True)
        self.search_edit.setMaximumWidth(300)
        toolbar.addWidget(self.search_edit)

        toolbar.addStretch()

        self.refresh_btn = QPushButton("Обновить")
        self.convert_btn = QPushButton("Конвертировать в XML")
        self.clear_btn = QPushButton("Очистить")
        self.export_btn = QPushButton("Экспорт XLSX")

        self.convert_btn.setStyleSheet("""
            QPushButton { color: white; background-color: #27AE60;
                border: none; padding: 8px 16px;
                border-radius: 5px; font-weight: bold}
            QPushButton:hover { background-color: #219A52}
        """)
        self.clear_btn.setStyleSheet("""
            QPushButton { color: white; background-color: #E74C3C;
                border: none; padding: 8px 16px;
                border-radius: 5px; font-weight: bold}
            QPushButton:hover { background-color: #C0392B}
        """)

        toolbar.addWidget(self.refresh_btn)
        toolbar.addWidget(self.convert_btn)
        toolbar.addWidget(self.clear_btn)
        toolbar.addWidget(self.export_btn)
        layout.addLayout(toolbar)

        self._model = DataViewTableModel(FIELD_KEYS, COLUMN_LABELS, self)
        self._proxy = MultiColumnFilterProxyModel(self)
        self._proxy.setSourceModel(self._model)

        self.table = QTableView()
        self.table.setModel(self._proxy)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setSortingEnabled(True)
        self.table.verticalHeader().setDefaultSectionSize(32)
        self.table.verticalHeader().setVisible(False)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.horizontalHeader().setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)

        self.status_label = QLabel("Записей: 0 | Выбрано: 0")

        layout.addWidget(self.table)
        layout.addWidget(self.status_label)

    def _connect_signals(self):
        self.table.customContextMenuRequested.connect(self._show_context_menu)
        self.table.selectionModel().selectionChanged.connect(self._update_status)
        self.table.horizontalHeader().customContextMenuRequested.connect(self._show_header_menu)
        self.search_edit.textChanged.connect(self._filter_table)
        self.refresh_btn.clicked.connect(self._load_all_data)
        self.convert_btn.clicked.connect(self.convert_to_xml)
        self.clear_btn.clicked.connect(self.clear_all_data)
        self.export_btn.clicked.connect(self._export_xlsx)
        self.table.doubleClicked.connect(self._on_item_double_click)

    def _load_all_data(self):
        rows = WorkersDataRepo.get_all()
        records = []
        for r in rows:
            records.append({
                'last_name': r['last_name'],
                'first_name': r['first_name'],
                'middle_name': r['middle_name'],
                'snils': r['snils'],
                'position': r['position'],
                'employer_inn': r['employer_inn'],
                'employer_title': r['employer_title'],
                'tc_inn': r['tc_inn'],
                'tc_title': r['tc_title'],
                'result': r['result'],
                'program': str(r['program']),
                'date': r['date'],
                'protocol': r['protocol'],
                'id': r['id'],
            })
        self._model.load_records(records)
        self._model.set_cell_color(9, QColor("#FFF0F0"))
        self._update_status()

    def get_existing_keys(self):
        return WorkersDataRepo.get_existing_keys()

    def _on_item_double_click(self, index):
        if not index.isValid():
            return
        col = index.column()
        if col == 3:
            snils = self._model.get_row_data(self._proxy.mapToSource(index).row()).get('snils', '').strip()
            if snils and snils.replace('-', '').replace(' ', '').isdigit():
                from PySide6.QtWidgets import QApplication
                QApplication.clipboard().setText(snils)
                from utils.toast import Toast
                Toast.info(self, f"СНИЛС скопирован: {snils}")

    def add_data(self, new_data, merge_mode=False):
        if merge_mode:
            WorkersDataRepo.clear()

        existing_keys = WorkersDataRepo.get_existing_keys()
        duplicates = 0
        to_add = []

        for row in new_data:
            key = (hash_for_search(row.get('snils', '')), str(row.get('program', '')), row.get('date', '') or '')
            if key in existing_keys:
                duplicates += 1
                continue
            existing_keys.add(key)
            to_add.append(row)

        if to_add:
            WorkersDataRepo.add_many(to_add)

        self._load_all_data()

        QMessageBox.information(
            self, "Загрузка завершена",
            f"Добавлено записей: {len(to_add)}\n"
            f"Пропущено дублей: {duplicates}"
        )

    def _filter_table(self, text):
        self._proxy.setFilterFixedString(text.strip())
        self._update_status()

    def _update_status(self):
        total = self._model.rowCount()
        visible = self._proxy.rowCount()
        selected_rows = len(self.table.selectionModel().selectedRows()) if self.table.selectionModel() else 0
        self.status_label.setText(f"Записей: {visible} | Выбрано: {selected_rows}")

    def _show_context_menu(self, position):
        idx = self.table.indexAt(position)
        if not idx.isValid():
            return

        menu = QMenu(self)
        menu.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)

        edit_action = menu.addAction("Редактировать")
        delete_action = menu.addAction("Удалить")
        duplicate_action = menu.addAction("Дублировать запись")

        action = menu.exec(self.table.mapToGlobal(position))
        if action == edit_action:
            self.edit_selected_row()
        elif action == delete_action:
            self.delete_selected_rows()
        elif action == duplicate_action:
            self._duplicate_selected_rows()

    def _show_header_menu(self, position):
        menu = QMenu(self)
        for i, label in enumerate(COLUMN_LABELS):
            clean = label.replace('\n', ' ')
            action = menu.addAction(clean)
            action.setCheckable(True)
            action.setChecked(not self.table.isColumnHidden(i))
            action.triggered.connect(lambda checked, col=i: self._toggle_column(col))
        menu.exec(self.table.horizontalHeader().mapToGlobal(position))

    def _toggle_column(self, col):
        self.table.setColumnHidden(col, not self.table.isColumnHidden(col))

    def _get_selected_source_rows(self) -> list:
        rows = set()
        for idx in self.table.selectionModel().selectedIndexes():
            rows.add(self._proxy.mapToSource(idx).row())
        return sorted(rows)

    def edit_selected_row(self):
        rows = self._get_selected_source_rows()
        if not rows:
            return
        row = rows[0]
        row_data = self._model.get_row_data(row)
        col_data = {}
        for col, key in enumerate(FIELD_KEYS):
            col_data[col] = row_data.get(key, '')
        col_data['id'] = row_data.get('id')
        dialog = EditDialog(col_data, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            new_data = dialog.get_data()
            record_id = col_data.get('id')
            if record_id:
                record = {}
                for col, key in enumerate(FIELD_KEYS):
                    record[key] = new_data.get(col, '')
                WorkersDataRepo.update(record_id, record)
                self._load_all_data()

    def delete_selected_rows(self):
        rows = self._get_selected_source_rows()
        if not rows:
            return
        count = len(rows)
        msg = "Вы уверены, что хотите удалить выбранные записи?" if count > 1 else "Вы уверены, что хотите удалить?"
        reply = QMessageBox.question(
            self, "Подтверждение", msg,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            for row in reversed(rows):
                record_id = self._model.get_record_id(row)
                if record_id:
                    WorkersDataRepo.delete(record_id)
            self._load_all_data()

    def _duplicate_selected_rows(self):
        rows = self._get_selected_source_rows()
        if not rows:
            return
        for row in rows:
            data = self._model.get_row_data(row)
            data.pop('id', None)
            WorkersDataRepo.add(data)
        self._load_all_data()

    def clear_all_data(self):
        reply = QMessageBox.question(
            self, "Подтверждение",
            "Очистить все данные? Данные удалятся безвозвратно.",
            QMessageBox.StandardButton.Ok | QMessageBox.StandardButton.Cancel
        )
        if reply == QMessageBox.StandardButton.Ok:
            WorkersDataRepo.clear()
            self._model.load_records([])
            self._update_status()

    def convert_to_xml(self):
        if self._model.rowCount() == 0:
            QMessageBox.warning(self, "Предупреждение", "Нет данных для конвертации")
            return

        from utils.app_paths import get_resource_dir, get_app_data_dir
        resource_dir = get_resource_dir()
        schema_dir = os.path.join(resource_dir, "schema")
        os.makedirs(schema_dir, exist_ok=True)
        xsd_files = [f for f in os.listdir(schema_dir) if f.endswith('.xsd')]

        if not xsd_files:
            QMessageBox.warning(self, "Предупреждение", "XSD отсутствует")
            return

        data_dir = get_app_data_dir()
        settings_file = os.path.join(data_dir, "org_settings.json")
        org_settings = {}
        if os.path.exists(settings_file):
            try:
                import json
                with open(settings_file, 'r', encoding='utf-8') as f:
                    wrapper = json.load(f)
                encrypted = wrapper.get('data', '')
                if encrypted:
                    org_settings = decrypt_data(encrypted)
                else:
                    org_settings = wrapper
            except (OSError, json.JSONDecodeError, ValueError) as e:
                logger.debug("Could not load org settings: %s", e)

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить XML", "", "XML Files (*.xml)"
        )

        if file_path:
            records = []
            for row in range(self._model.rowCount()):
                records.append(self._model.get_row_data(row))

            success, message = export_to_xml(records, file_path, org_settings)
            if success:
                QMessageBox.information(self, "Успех", message)
            else:
                QMessageBox.warning(self, "Ошибка", message)

    def _export_xlsx(self):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        total_count = self._model.rowCount()
        if total_count == 0:
            QMessageBox.warning(self, "Предупреждение", "Нет данных для экспорта")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить XLSX", "", "Excel Files (*.xlsx)"
        )
        if not file_path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Данные"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4169E1", end_color="4169E1", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col, label in enumerate(COLUMN_LABELS, 1):
            cell = ws.cell(row=1, column=col, value=label.replace('\n', ' '))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        row_num = 2
        for src_row in range(total_count):
            rec = self._model.get_row_data(src_row)
            for col in range(len(COLUMN_LABELS)):
                key = FIELD_KEYS[col] if col < len(FIELD_KEYS) else ''
                value = rec.get(key, '')
                cell = ws.cell(row=row_num, column=col + 1, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            row_num += 1

        exported = row_num - 2
        for col in range(1, len(COLUMN_LABELS) + 1):
            max_len = len(COLUMN_LABELS[col - 1].replace('\n', ' '))
            for row in range(2, row_num):
                val = ws.cell(row=row, column=col).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_len + 3, 50)

        wb.save(file_path)
        QMessageBox.information(self, "Успех", f"Экспортировано записей: {exported}")


class EditDialog(BaseDialog):
    def __init__(self, row_data: dict, parent=None):
        super().__init__(parent, title="Редактирование данных", min_width=520, min_height=520)

        self.row_data = row_data
        bl = self.body_layout()
        form_layout = QFormLayout()

        self.fields: dict = {}
        self.field_widgets: dict = {}
        field_names = [
            ("Фамилия", 0), ("Имя", 1), ("Отчество", 2), ("СНИЛС", 3),
            ("Должность", 4), ("ИНН Заказчика", 5), ("Наименование ЮЛ заказчика", 6),
            ("ИНН УЦ", 7), ("Наименование УЦ", 8), ("Результат", 9),
            ("№ программы", 10), ("Дата", 11), ("№ протокола", 12)
        ]

        from utils.field_validators import validate_required, validate_name, validate_snils, validate_program_id, validate_date
        from utils.field_validators import ValidatedLineEdit

        for label, col_idx in field_names:
            if col_idx == 9:
                combo = QComboBox()
                combo.addItems(["Удовлетворительно", "Неудовлетворительно"])
                current_value = row_data.get(col_idx, "Удовлетворительно")
                idx = combo.findText(current_value)
                if idx >= 0:
                    combo.setCurrentIndex(idx)
                form_layout.addRow(label, combo)
                self.fields[col_idx] = combo
                self.field_widgets[col_idx] = {'widget': combo, 'label': label}
            elif col_idx in (0, 1, 4):
                le = ValidatedLineEdit(
                    row_data.get(col_idx, ""),
                    validator=lambda t, idx=col_idx: validate_required(t, field_names[idx][0]) or validate_name(t) if t else None
                )
                self.register_field(le)
                form_layout.addRow(label, le)
                self.fields[col_idx] = le
                self.field_widgets[col_idx] = {'widget': le, 'label': label}
            elif col_idx == 2:
                le = ValidatedLineEdit(row_data.get(col_idx, ""), validator=validate_name)
                form_layout.addRow(label, le)
                self.fields[col_idx] = le
                self.field_widgets[col_idx] = {'widget': le, 'label': label}
            elif col_idx == 3:
                le = ValidatedLineEdit(row_data.get(col_idx, ""), validator=validate_snils)
                form_layout.addRow(label, le)
                self.fields[col_idx] = le
                self.field_widgets[col_idx] = {'widget': le, 'label': label}
            elif col_idx == 10:
                le = ValidatedLineEdit(row_data.get(col_idx, ""), validator=validate_program_id)
                form_layout.addRow(label, le)
                self.fields[col_idx] = le
                self.field_widgets[col_idx] = {'widget': le, 'label': label}
            elif col_idx == 11:
                le = ValidatedLineEdit(row_data.get(col_idx, ""), validator=validate_date)
                form_layout.addRow(label, le)
                self.fields[col_idx] = le
                self.field_widgets[col_idx] = {'widget': le, 'label': label}
            else:
                line_edit = QLineEdit()
                line_edit.setText(row_data.get(col_idx, ""))
                form_layout.addRow(label, line_edit)
                self.fields[col_idx] = line_edit
                self.field_widgets[col_idx] = {'widget': line_edit, 'label': label}

        bl.addLayout(form_layout)

        btn_layout = QHBoxLayout()
        save_btn = QPushButton("Сохранить")
        save_btn.setObjectName("dialogPrimaryBtn")
        save_btn.clicked.connect(self.validate_and_save)
        cancel_btn = QPushButton("Отмена")
        cancel_btn.setObjectName("dialogDangerBtn")
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(save_btn)
        btn_layout.addWidget(cancel_btn)
        bl.addLayout(btn_layout)

    def _set_error(self, col_idx: int):
        info = self.field_widgets.get(col_idx)
        if info:
            w = info['widget']
            if hasattr(w, 'set_invalid'):
                w.set_invalid("Проверьте значение")
            else:
                w.setStyleSheet("border: 2px solid #E74C3C;")

    def _clear_errors(self):
        for info in self.field_widgets.values():
            w = info['widget']
            if hasattr(w, 'clear_validation'):
                w.clear_validation()
            else:
                w.setStyleSheet("")

    def validate_and_save(self):
        self._clear_errors()

        values = {}
        for col_idx, info in self.field_widgets.items():
            if isinstance(info['widget'], QComboBox):
                values[col_idx] = info['widget'].currentText()
            else:
                values[col_idx] = info['widget'].text().strip()

        import unicodedata
        snils_raw = values.get(3, '')
        snils_clean = ''.join(c for c in snils_raw if unicodedata.category(c) != 'Zs')
        snils_clean = snils_clean.replace('-', '')
        if snils_clean and (not snils_clean.isdigit() or len(snils_clean) != 11):
            self._set_error(3)
            from utils.error_utils import show_error_dialog
            show_error_dialog(self, "Ошибка", "СНИЛС должен содержать 11 цифр")
            return

        date_val = values.get(11, '').strip()
        if date_val:
            clean = date_val.replace('.', '').replace('-', '')
            if not clean.isdigit() or len(clean) != 8:
                self._set_error(11)
                from utils.error_utils import show_error_dialog
                show_error_dialog(self, "Ошибка", "Дата должна быть в формате ДД.ММ.ГГГГ")
                return
            try:
                from datetime import datetime
                dt = datetime.strptime(clean, "%d%m%Y")
                if dt.date() > datetime.now().date():
                    self._set_error(11)
                    from utils.error_utils import show_error_dialog
                    show_error_dialog(self, "Ошибка", "Дата не может быть больше текущей")
                    return
            except ValueError:
                self._set_error(11)
                from utils.error_utils import show_error_dialog
                show_error_dialog(self, "Ошибка", "Дата некорректна")
                return

        self.accept()

    def get_data(self) -> dict:
        data = {}
        for col_idx, widget in self.fields.items():
            if isinstance(widget, QComboBox):
                data[col_idx] = widget.currentText()
            else:
                data[col_idx] = widget.text()
        return data
