import os
import json
import logging
from datetime import datetime
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableView,
    QPushButton, QLabel, QLineEdit, QComboBox, QHeaderView, QAbstractItemView,
    QMessageBox, QFileDialog, QDateEdit, QFrame, QMenu, QInputDialog
)
from PySide6.QtCore import Qt, QDate, QCoreApplication
from PySide6.QtGui import QColor, QFont, QBrush, QIcon
from db.exam_journal_repo import JournalRecord
from utils.table_models import ExamJournalTableModel, JOURNAL_FIELD_NAMES
from utils.error_utils import safe_message_box

logger = logging.getLogger(__name__)


COL_STATUS = 13
COL_SETID = 11


class ExamJournalTab(QWidget):
    def __init__(self, journal_manager, data_dir=None):
        super().__init__()
        self.journal = journal_manager
        self.data_dir = data_dir
        self.last_save_path = self._load_last_save_path() if data_dir else ""

        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)

        layout.addWidget(self._create_filter_panel())
        layout.addWidget(self._create_toolbar())
        self.table = self._create_table()
        layout.addWidget(self.table)
        self.status_label = QLabel("Записей: 0 | Найдено: 0")
        layout.addWidget(self.status_label)

        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._show_context_menu)

        self.refresh_journal()

    # ── Persistence ───────────────────────────────────────────

    def _load_last_save_path(self):
        if not self.data_dir:
            return ""
        fp = os.path.join(self.data_dir, "journal_settings.json")
        if os.path.exists(fp):
            try:
                with open(fp, 'r', encoding='utf-8') as f:
                    return json.load(f).get('last_save_path', '')
            except (OSError, json.JSONDecodeError) as e:
                logger.debug("Cannot load last_save_path: %s", e)
        return ""

    def _save_last_save_path(self, path):
        if not self.data_dir:
            return
        fp = os.path.join(self.data_dir, "journal_settings.json")
        try:
            settings = {}
            if os.path.exists(fp):
                with open(fp, 'r', encoding='utf-8') as f:
                    settings = json.load(f)
            settings['last_save_path'] = path
            with open(fp, 'w', encoding='utf-8') as f:
                json.dump(settings, f, ensure_ascii=False, indent=2)
        except (OSError, json.JSONDecodeError) as e:
            logger.error("Cannot save last_save_path: %s", e, exc_info=True)

    # ── Filter panel ──────────────────────────────────────────

    def _create_filter_panel(self):
        group = QFrame()
        group.setFrameShape(QFrame.Shape.StyledPanel)

        layout = QVBoxLayout(group)
        layout.setSpacing(8)
        layout.setContentsMargins(10, 8, 10, 8)

        row1 = QHBoxLayout()
        row1.setSpacing(10)

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Поиск: Фамилия, Имя или СНИЛС")
        self.search_input.setClearButtonEnabled(True)
        self.search_input.textChanged.connect(self._apply_filters)
        row1.addWidget(QLabel("Поиск:"))
        row1.addWidget(self.search_input)

        self.setid_combo = QComboBox()
        self.setid_combo.setMinimumWidth(220)
        self.setid_combo.addItem("Все")
        self.setid_combo.currentTextChanged.connect(self._apply_filters)
        row1.addWidget(QLabel("SetId:"))
        row1.addWidget(self.setid_combo)

        self.status_combo = QComboBox()
        self.status_combo.addItems(["Все", "ожидает", "получен"])
        self.status_combo.currentTextChanged.connect(self._apply_filters)
        row1.addWidget(QLabel("Статус:"))
        row1.addWidget(self.status_combo)

        row1.addSpacing(8)
        self.date_from = QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_from.setDisplayFormat("dd.MM.yyyy")
        self.date_from.setDate(QDate(2023, 1, 1))
        self.date_from.dateChanged.connect(self._apply_filters)
        row1.addWidget(QLabel("Дата с:"))
        row1.addWidget(self.date_from)

        self.date_to = QDateEdit()
        self.date_to.setCalendarPopup(True)
        self.date_to.setDisplayFormat("dd.MM.yyyy")
        self.date_to.setDate(self.date_to.maximumDate())
        self.date_to.dateChanged.connect(self._apply_filters)
        row1.addWidget(QLabel("по:"))
        row1.addWidget(self.date_to)

        reset_btn = QPushButton("Сбросить фильтры")
        reset_btn.clicked.connect(self._reset_filters)
        row1.addWidget(reset_btn)
        row1.addStretch()
        layout.addLayout(row1)

        return group

    # ── Toolbar ───────────────────────────────────────────────

    def _create_toolbar(self):
        bar = QFrame()
        row = QHBoxLayout(bar)
        row.setContentsMargins(0, 4, 0, 4)
        row.setSpacing(8)

        exp = QPushButton("Экспорт в XLSX")
        exp.clicked.connect(self._export_to_xlsx)
        row.addWidget(exp)

        imp = QPushButton("Импорт из Excel")
        imp.clicked.connect(self._import_from_excel)
        row.addWidget(imp)

        tmpl = QPushButton("Шаблон журнала")
        tmpl.clicked.connect(self._create_journal_template)
        row.addWidget(tmpl)

        prt = QPushButton("Печать протоколов")
        prt.clicked.connect(self._print_protocol)
        row.addWidget(prt)

        upd = QPushButton("Обновить по SetId")
        upd.clicked.connect(self._update_selected_by_setid)
        row.addWidget(upd)

        self.delete_btn = QPushButton("Удалить выбранное")
        self.delete_btn.setStyleSheet("""
            QPushButton { color: white; background-color: #E74C3C;
                border: none; padding: 7px 16px;
                border-radius: 5px; font-weight: bold}
            QPushButton:hover { background-color: #C0392B}
        """)
        self.delete_btn.clicked.connect(self._delete_selected)
        row.addWidget(self.delete_btn)

        row.addStretch()
        return bar

    # ── Table ─────────────────────────────────────────────────

    def _create_table(self):
        self._model = ExamJournalTableModel(self)

        table = QTableView()
        table.setModel(self._model)
        table.setAlternatingRowColors(True)
        table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        table.horizontalHeader().setStretchLastSection(False)
        table.verticalHeader().setDefaultSectionSize(30)
        table.verticalHeader().setVisible(False)
        table.setSortingEnabled(True)

        # Hide the 15th column (uuid column) - no, we now use UserRole
        # Set column widths matching old layout
        for col, w in enumerate([100, 100, 100, 100, 100, 120, 100, 70, 280, 100, 100, 120, 120, 90]):
            table.setColumnWidth(col, w)

        return table

    # ── Refresh ───────────────────────────────────────────────

    def refresh_journal(self):
        filtered = self._get_filtered_records()
        self._model.load_records(filtered)
        total = len(self.journal.get_all_records())
        shown = len(filtered)
        self.status_label.setText(
            f"Записей: {total} | Найдено: {shown}"
        )
        self._update_setid_combo()

    def _get_filtered_records(self):
        query = self.search_input.text()
        set_id = self.setid_combo.currentText()
        if set_id == "Все":
            set_id = ""
        status = self.status_combo.currentText()
        if status == "Все":
            status = "all"
        else:
            status = "pending" if status == "ожидает" else "received"

        date_from = self.date_from.date().toString("dd.MM.yyyy") if self.date_from.date() > QDate(2023, 1, 1) else ""
        date_to = self.date_to.date().toString("dd.MM.yyyy") if self.date_to.date() < self.date_to.maximumDate() else ""

        return self.journal.search(
            query=query, set_id=set_id,
            status=status, date_from=date_from, date_to=date_to
        )

    def _update_setid_combo(self):
        current = self.setid_combo.currentData()
        self.setid_combo.blockSignals(True)
        self.setid_combo.clear()
        self.setid_combo.addItem("Все")
        for sid in self.journal.get_unique_set_ids():
            short = sid[:22] + "..." if len(sid) > 22 else sid
            self.setid_combo.addItem(short, sid)
        idx = self.setid_combo.findData(current)
        self.setid_combo.setCurrentIndex(idx if idx >= 0 else 0)
        self.setid_combo.blockSignals(False)

    # ── Filters ───────────────────────────────────────────────

    def _apply_filters(self):
        self.refresh_journal()

    def _reset_filters(self):
        self.search_input.clear()
        self.setid_combo.setCurrentIndex(0)
        self.status_combo.setCurrentIndex(0)
        self.date_from.setDate(QDate(2023, 1, 1))
        self.date_to.setDate(self.date_to.maximumDate())
        self.refresh_journal()

    # ── Context menu ──────────────────────────────────────────

    def _show_context_menu(self, position):
        if not self.table.selectedIndexes():
            idx = self.table.indexAt(position)
            if idx.isValid():
                self.table.selectRow(idx.row())
        menu = QMenu(self)
        menu.addAction("Удалить", self._delete_selected)
        menu.addAction("Обновить по SetId", self._update_selected_by_setid)
        menu.exec(self.table.mapToGlobal(position))

    # ── Update by SetId ───────────────────────────────────────

    def _update_selected_by_setid(self):
        selected = sorted(set(it.row() for it in self.table.selectedIndexes()))
        if not selected:
            QMessageBox.information(self, "Информация", "Выберите записи для обновления")
            return

        set_ids = set()
        for row in selected:
            rec = self._model.get_record(row)
            if rec and rec.set_id:
                set_ids.add(rec.set_id)

        if not set_ids:
            QMessageBox.warning(self, "Ошибка", "У выбранных записей нет SetId")
            return

        try:
            from api.mintrud_api import get_by_set_id
        except ImportError:
            QMessageBox.warning(self, "Ошибка", "API модуль недоступен")
            return

        from utils.proxy_manager import load_proxy_settings
        from utils.app_paths import get_app_data_dir
        proxy = load_proxy_settings(get_app_data_dir())

        from api.mintrud_api import load_api_key
        api_key = load_api_key(get_app_data_dir())

        if not api_key or len(api_key) != 32:
            QMessageBox.warning(self, "Ошибка", "API ключ не настроен")
            return

        updated_total = 0
        for sid in set_ids:
            result = get_by_set_id(api_key, sid, proxy_settings=proxy)
            if result.get("success"):
                records = result.get("records", [])
                if records:
                    base_no_map = {}
                    for rec in records:
                        snils = rec.get('Snils', '').replace('-', '').replace(' ', '')
                        base_no = rec.get('baseNo', '')
                        if snils:
                            base_no_map[snils] = base_no
                    if base_no_map:
                        cnt = self.journal.update_base_no_by_set_id(sid, base_no_map)
                        updated_total += cnt
            else:
                logger.warning(f"SetId {sid} query failed: {result.get('error')}")

        self.refresh_journal()
        if updated_total:
            QMessageBox.information(self, "Успех", f"Обновлено записей: {updated_total}")
        else:
            QMessageBox.information(self, "Информация", "Нет данных для обновления")

    # ── Export XLSX ───────────────────────────────────────────

    def _export_to_xlsx(self):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        records = self._get_filtered_records()
        if not records:
            QMessageBox.information(self, "Информация", "Нет данных для экспорта")
            return

        if self.last_save_path and os.path.exists(os.path.dirname(self.last_save_path)):
            default = os.path.join(os.path.dirname(self.last_save_path), "Журнал_проверки_знаний.xlsx")
        else:
            default = "Журнал_проверки_знаний.xlsx"

        fp, _ = QFileDialog.getSaveFileName(self, "Сохранить XLSX", default, "Excel Files (*.xlsx)")
        if not fp:
            return
        self._save_last_save_path(fp)

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Журнал"
            ws.append(JOURNAL_FIELD_NAMES)

            hf = Font(bold=True, color="FFFFFF")
            hfill = PatternFill(start_color="4169E1", end_color="4169E1", fill_type="solid")
            for cell in ws[1]:
                cell.font = hf
                cell.fill = hfill
                cell.alignment = Alignment(horizontal="center")

            for rec in records:
                ws.append([
                    rec.protocol, rec.exam_date.split()[0] if rec.exam_date else "",
                    rec.last_name, rec.first_name, rec.middle_name, rec.snils,
                    rec.base_no, rec.program_id, rec.program_title, rec.position,
                    rec.result, rec.set_id, rec.send_date.split()[0] if rec.send_date else "",
                    "получен" if rec.status == "received" else "ожидает"
                ])

            for col in ws.columns:
                mx = max((len(str(c.value or "")) for c in col), default=0)
                ws.column_dimensions[col[0].column_letter].width = min(mx + 2, 50)
            wb.save(fp)
            safe_message_box(self, QMessageBox.Icon.Information, "Успех", f"Журнал сохранён:\n{fp}")
        except Exception as e:
            logger.exception("Journal export error")
            safe_message_box(self, QMessageBox.Icon.Warning, "Ошибка", f"Ошибка экспорта: {e}")

    # ── Import from Excel ─────────────────────────────────────

    def _import_from_excel(self):
        from openpyxl import load_workbook
        fp, _ = QFileDialog.getOpenFileName(self, "Выберите Excel файл", "", "Excel Files (*.xlsx *.xls)")
        if not fp:
            return
        try:
            wb = load_workbook(fp)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]

            journal_headers = [
                "№ протокола", "Дата экзамена", "Фамилия", "Имя", "Отчество", "СНИЛС",
                "Рег. номер", "№ программы", "Название программы", "Должность",
                "Результат", "SetId", "Дата отправки", "Статус"
            ]
            api_headers = [
                "Номер записи в реестре", "Фамилия", "Имя", "Отчество",
                "СНИЛС", "Номер программы", "Название программы",
                "Номер протокола", "Дата"
            ]
            import_headers = [
                "last_name", "first_name", "middle_name", "snils", "position",
                "program_id", "program_title", "exam_date", "protocol", "result",
                "set_id", "base_no", "status"
            ]

            is_journal = all(h in headers for h in journal_headers)
            is_api = all(h in headers for h in api_headers)
            is_import = all(h in headers for h in import_headers)

            if not (is_journal or is_api or is_import):
                QMessageBox.warning(self, "Ошибка", "Неподдерживаемый формат файла")
                return

            ci = {h: i for i, h in enumerate(headers)}
            records = []
            errors = []

            for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                try:
                    if not any(c is not None for c in row):
                        continue

                    if is_journal:
                        proto = str(row[ci.get("№ протокола", 0)] or "").strip()
                        exam = str(row[ci.get("Дата экзамена", 1)] or "").strip()
                        ln = str(row[ci.get("Фамилия", 2)] or "").strip()
                        fn = str(row[ci.get("Имя", 3)] or "").strip()
                        mn = str(row[ci.get("Отчество", 4)] or "").strip()
                        sn = str(row[ci.get("СНИЛС", 5)] or "").strip()
                        bn = str(row[ci.get("Рег. номер", 6)] or "").strip()
                        pid = str(row[ci.get("№ программы", 7)] or "").strip()
                        pt = str(row[ci.get("Название программы", 8)] or "").strip()
                        pos = str(row[ci.get("Должность", 9)] or "").strip()
                        res = str(row[ci.get("Результат", 10)] or "").strip()
                        sid = str(row[ci.get("SetId", 11)] or "").strip()
                        sd = str(row[ci.get("Дата отправки", 12)] or "").strip()
                        st = str(row[ci.get("Статус", 13)] or "").lower()
                        if not all([ln, fn, mn, sn, pid, pt, proto, exam]):
                            errors.append(f"Строка {r_idx}: заполните обязательные поля")
                            continue
                        s_digits = ''.join(filter(str.isdigit, sn))
                        if len(s_digits) != 11:
                            errors.append(f"Строка {r_idx}: СНИЛС должен содержать 11 цифр")
                            continue
                        sn_fmt = f"{s_digits[:3]}-{s_digits[3:6]}-{s_digits[6:9]} {s_digits[9:]}"
                        status = "received" if "получен" in st or "received" in st else "pending"
                        res = res or "Удовлетворительно"

                    elif is_api:
                        bn = str(row[ci.get("Номер записи в реестре", 0)] or "").strip()
                        ln = str(row[ci.get("Фамилия", 1)] or "").strip()
                        fn = str(row[ci.get("Имя", 2)] or "").strip()
                        mn = str(row[ci.get("Отчество", 3)] or "").strip()
                        sn = str(row[ci.get("СНИЛС", 4)] or "").strip()
                        pid = str(row[ci.get("Номер программы", 5)] or "").strip()
                        pt = str(row[ci.get("Название программы", 6)] or "").strip()
                        proto = str(row[ci.get("Номер протокола", 7)] or "").strip()
                        exam = str(row[ci.get("Дата", 8)] or "").strip()
                        if not all([ln, fn, mn, sn, pid, pt, proto, exam]):
                            errors.append(f"Строка {r_idx}: заполните обязательные поля")
                            continue
                        s_digits = ''.join(filter(str.isdigit, sn))
                        if len(s_digits) != 11:
                            errors.append(f"Строка {r_idx}: СНИЛС должен содержать 11 цифр")
                            continue
                        sn_fmt = f"{s_digits[:3]}-{s_digits[3:6]}-{s_digits[6:9]} {s_digits[9:]}"
                        status = "received" if bn else "pending"
                        res = "Удовлетворительно"
                        sd = sid = pos = ""

                    else:
                        ln = str(row[ci.get("last_name", 0)] or "").strip()
                        fn = str(row[ci.get("first_name", 1)] or "").strip()
                        mn = str(row[ci.get("middle_name", 2)] or "").strip()
                        sn = str(row[ci.get("snils", 3)] or "").strip()
                        pos = str(row[ci.get("position", 4)] or "").strip()
                        pid = str(row[ci.get("program_id", 5)] or "").strip()
                        pt = str(row[ci.get("program_title", 6)] or "").strip()
                        exam = str(row[ci.get("exam_date", 7)] or "").strip()
                        proto = str(row[ci.get("protocol", 8)] or "").strip()
                        res = str(row[ci.get("result", 9)] or "").strip()
                        sid = str(row[ci.get("set_id", 10)] or "").strip()
                        bn = str(row[ci.get("base_no", 11)] or "").strip()
                        st = str(row[ci.get("status", 12)] or "").lower()
                        if not all([ln, fn, mn, sn, pid, pt, proto, exam]):
                            errors.append(f"Строка {r_idx}: заполните обязательные поля")
                            continue
                        s_digits = ''.join(filter(str.isdigit, sn))
                        if len(s_digits) != 11:
                            errors.append(f"Строка {r_idx}: СНИЛС должен содержать 11 цифр")
                            continue
                        sn_fmt = f"{s_digits[:3]}-{s_digits[3:6]}-{s_digits[6:9]} {s_digits[9:]}"
                        status = "received" if "получен" in st or "received" in st else "pending"
                        res = res or "Удовлетворительно"
                        sd = ""

                    import uuid
                    records.append(JournalRecord(
                        uuid=str(uuid.uuid4()), send_date=sd, set_id=sid, xml_file="",
                        last_name=ln, first_name=fn, middle_name=mn, snils=sn_fmt,
                        position=pos, program_id=pid, program_title=pt,
                        exam_date=exam, protocol=proto, result=res, base_no=bn, status=status
                    ))
                except (ValueError, TypeError, KeyError) as e:
                    logger.exception("Row %d import error", r_idx)
                    errors.append(f"Строка {r_idx}: {e}")

            if errors:
                logger.error("Import errors: %s", errors)
                QMessageBox.warning(self, "Ошибки импорта",
                                    "Обнаружены ошибки:\n" + "\n".join(errors[:10]) +
                                    (f"\n... и ещё {len(errors)-10}" if len(errors) > 10 else ""))
            if not records:
                QMessageBox.information(self, "Информация", "Нет данных для импорта")
                return
            self.journal.add_journal_records_directly(records)
            QMessageBox.information(self, "Успех", f"Импортировано {len(records)} записей")
            self.refresh_journal()
        except (OSError, ValueError, TypeError) as e:
            logger.error("Import failed: %s", e, exc_info=True)
            QMessageBox.critical(self, "Ошибка", f"Не удалось прочитать файл:\n{e}")

    # ── Create template ───────────────────────────────────────

    def _create_journal_template(self):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        fp, _ = QFileDialog.getSaveFileName(self, "Сохранить шаблон", "Шаблон_журнала.xlsx", "Excel Files (*.xlsx)")
        if not fp:
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Журнал"
            ws.append(JOURNAL_FIELD_NAMES)
            hf = Font(bold=True, color="FFFFFF")
            hfill = PatternFill(start_color="4169E1", end_color="4169E1", fill_type="solid")
            for cell in ws[1]:
                cell.font = hf
                cell.fill = hfill
                cell.alignment = Alignment(horizontal="center")
            ws.append(["1", "01.01.2026", "Иванов", "Иван", "Иванович", "123-456-789 00",
                       "123456", "1", "Оказание первой помощи пострадавшим", "Инженер",
                       "Удовлетворительно", "SET123456", "01.01.2026", "получен"])
            for col in ws.columns:
                mx = max((len(str(c.value or "")) for c in col), default=0)
                ws.column_dimensions[col[0].column_letter].width = min(mx + 2, 50)
            wb.save(fp)
            QMessageBox.information(self, "Успех", f"Шаблон создан:\n{fp}")
        except Exception as e:
            logger.exception("Template creation error")
            QMessageBox.warning(self, "Ошибка", f"Ошибка: {e}")

    # ── Print protocol ────────────────────────────────────────

    def _print_protocol(self):
        from exporters.protocol_exporter import ProtocolExporter

        records = self._get_filtered_records()
        if not records:
            QMessageBox.information(self, "Информация", "Нет данных")
            return

        protocol_numbers = sorted(set(r.protocol for r in records if r.protocol))
        if not protocol_numbers:
            QMessageBox.warning(self, "Ошибка", "Нет номеров протоколов в записях")
            return

        choices = ["Все"] + protocol_numbers
        if len(choices) == 2:
            proto_num = choices[0]
        else:
            proto_num, ok = QInputDialog.getItem(self, "Номер протокола", "Выберите номер:", choices, 0, False)
            if not ok or not proto_num:
                return

        if proto_num == "Все":
            default = "Протоколы.docx"
        else:
            ed = ""
            for r in records:
                if r.protocol == proto_num and r.exam_date:
                    try:
                        dt = datetime.strptime(r.exam_date.split()[0], "%d.%m.%Y")
                        ed = " от " + dt.strftime("%d-%m-%Y")
                    except ValueError:
                        ed = ""
                    break
            default = f"Протокол {proto_num}{ed}.docx"

        base_dir = os.path.dirname(self.last_save_path) if self.last_save_path else ""
        fp, _ = QFileDialog.getSaveFileName(self, "Сохранить протокол",
                                            os.path.join(base_dir, default) if base_dir else default,
                                            "Word Files (*.docx)")
        if not fp:
            return
        self._save_last_save_path(fp)

        from utils.app_paths import get_app_data_dir, get_resource_dir
        data_dir = get_app_data_dir()
        tmpl = os.path.join(get_resource_dir(), "templates", "Protokol_proverki_znanii_OT.docx")

        if proto_num == "Все":
            save_dir = QFileDialog.getExistingDirectory(self, "Папка для протоколов",
                                                        os.path.dirname(fp) if fp else "")
            if not save_dir:
                return
            saved = 0
            for pn in protocol_numbers:
                sel = [r for r in records if r.protocol == pn]
                if not sel:
                    continue
                ed = ""
                for r in sel:
                    if r.exam_date:
                        try:
                            dt = datetime.strptime(r.exam_date.split()[0], "%d.%m.%Y")
                            ed = "_" + dt.strftime("%d-%m-%Y")
                        except ValueError:
                            pass
                        break
                ok, _ = ProtocolExporter.export_protocol(sel, os.path.join(save_dir, f"Протокол {pn}{ed}.docx"), tmpl, data_dir)
                if ok:
                    saved += 1
            QMessageBox.information(self, "Успех", f"Сохранено протоколов: {saved}")
        else:
            sel = [r for r in records if r.protocol == proto_num]
            if not sel:
                QMessageBox.warning(self, "Ошибка", "Нет данных для выбранного протокола")
                return
            ok, msg = ProtocolExporter.export_protocol(sel, fp, tmpl, data_dir)
            (QMessageBox.information if ok else QMessageBox.warning)(self, "Успех" if ok else "Ошибка", msg)

    # ── Delete selected ───────────────────────────────────────

    def _delete_selected(self):
        rows = sorted(set(it.row() for it in self.table.selectedIndexes()))
        if not rows:
            QMessageBox.information(self, "Информация", "Выберите записи для удаления")
            return
        if QMessageBox.question(self, "Подтверждение",
                                f"Удалить выбранные записи ({len(rows)} шт.)?\nДанные удалятся безвозвратно.",
                                QMessageBox.StandardButton.Ok | QMessageBox.StandardButton.Cancel) != QMessageBox.StandardButton.Ok:
            return

        uuids = []
        for row in rows:
            rec = self._model.get_record(row)
            if rec:
                uuids.append(rec.uuid)
        self.journal.delete_by_uuid(uuids)
        self.refresh_journal()
        QMessageBox.information(self, "Успех", "Записи удалены")

    # ── Public API ────────────────────────────────────────────

    def add_records_to_journal(self, records_data, set_id, xml_file):
        count = self.journal.add_records(records_data, set_id, xml_file)
        self.refresh_journal()
        return count

    def update_base_no(self, set_id, base_no_map):
        count = self.journal.update_base_no_by_set_id(set_id, base_no_map)
        if count > 0:
            self.refresh_journal()
        return count
