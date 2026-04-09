"""
Обработчик Excel файлов - создание шаблонов и импорт данных
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from typing import List, Tuple
import os


class ExcelHandler:
    """Класс для работы с Excel файлами"""

    COLUMNS = [
        'Фамилия', 'Имя', 'Отчество', 'СНИЛС', 'Должность',
        'ИНН Заказчика', 'Наименование ЮЛ Заказчика', 'ИНН УЦ',
        'Наименование УЦ', 'Результат', '№ программы', 'Дата', '№ протокола'
    ]

    def create_template(self, filepath: str) -> bool:
        """Создание шаблона Excel файла"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Данные"

            # Заголовки
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            for col, header in enumerate(self.COLUMNS, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Пример данных
            example_data = [
                ['Иванов', 'Иван', 'Иванович', '12345678901', 'Инженер',
                 '7701234567', 'ООО "Заказчик"', '7709876543',
                 'Учебный центр', 'Удовлетворительно', '1,2', '01.12.2024', '123-П']
            ]

            for row_idx, row_data in enumerate(example_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Автоширина колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(filepath)
            return True
        except Exception as e:
            print(f"Ошибка создания шаблона: {e}")
            return False

    def import_from_excel(self, filepath: str) -> Tuple[List[dict], List[str]]:
        """
        Импорт данных из Excel файла
        Возвращает кортеж (список записей, список ошибок)
        """
        records = []
        errors = []

        try:
            from openpyxl import load_workbook
            wb = load_workbook(filename=filepath, read_only=True)
            ws = wb.active

            # Чтение заголовков
            headers = []
            for cell in ws[1]:
                headers.append(cell.value.strip() if cell.value else '')

            # Проверка наличия необходимых колонок
            required_columns = set(self.COLUMNS)
            available_columns = set(headers)
            missing = required_columns - available_columns
            if missing:
                errors.append(f"Отсутствуют колонки: {', '.join(missing)}")
                return records, errors

            # Чтение данных
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                try:
                    record = {}
                    for col_idx, value in enumerate(row):
                        if col_idx < len(headers):
                            header = headers[col_idx]
                            if header in self.COLUMNS:
                                record[header] = str(value).strip() if value is not None else ''

                    # Простая валидация
                    if not record.get('Фамилия') or not record.get('Имя'):
                        errors.append(f"Строка {row_idx}: не заполнены обязательные поля")
                        continue

                    records.append(record)
                except Exception as e:
                    errors.append(f"Строка {row_idx}: {str(e)}")

            wb.close()

        except Exception as e:
            errors.append(f"Ошибка чтения файла: {str(e)}")

        return records, errors