"""
Парсер ответов от сервера Минтруда
"""

from xml.etree import ElementTree as ET
from openpyxl import Workbook
from typing import List, Dict


class ResponseParser:
    """Парсер XML ответов от Минтруда"""

    def parse_to_excel(self, xml_content: str, output_path: str) -> tuple[bool, str]:
        """
        Парсинг XML ответа и сохранение в Excel
        Возвращает кортеж (успех, сообщение)
        """
        try:
            # Парсим XML
            root = ET.fromstring(xml_content)

            # Находим все записи - поддерживаем оба формата
            records = []

            # Формат 1: RegistryRecord (прямой ответ от send_xml)
            for registry_record in root.findall('.//RegistryRecord'):
                record = self._parse_record(registry_record)
                if record:
                    records.append(record)

            # Формат 2: EducatedPerson (ответ от GetEducatedPersonXML по SetId/SNILS)
            for educated_person in root.findall('.//EducatedPerson'):
                record = self._parse_educated_person(educated_person)
                if record:
                    records.append(record)

            if not records:
                return False, "Нет данных для экспорта"

            # Создаем Excel файл
            wb = Workbook()
            ws = wb.active
            ws.title = "Данные из реестра"

            # Заголовки
            headers = list(records[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            # Данные
            for row_idx, record in enumerate(records, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=record.get(header, ''))

            # Автоширина колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(output_path)

            return True, f"Экспортировано {len(records)} записей"

        except ET.ParseError as e:
            return False, f"Ошибка парсинга XML: {str(e)}"
        except Exception as e:
            return False, f"Ошибка: {str(e)}"

    def _parse_record(self, registry_record: ET.Element) -> Dict[str, str]:
        """Парсинг одной записи RegistryRecord"""
        record = {}

        # Worker данные
        worker = registry_record.find('Worker')
        if worker is not None:
            record['Фамилия'] = self._get_text(worker, 'LastName')
            record['Имя'] = self._get_text(worker, 'FirstName')
            record['Отчество'] = self._get_text(worker, 'MiddleName')
            record['СНИЛС'] = self._get_text(worker, 'Snils')
            record['Должность'] = self._get_text(worker, 'Position')
            record['ИНН Заказчика'] = self._get_text(worker, 'EmployerInn')
            record['Наименование ЮЛ Заказчика'] = self._get_text(worker, 'EmployerTitle')

        # Organization данные
        organization = registry_record.find('Organization')
        if organization is not None:
            record['ИНН УЦ'] = self._get_text(organization, 'Inn')
            record['Наименование УЦ'] = self._get_text(organization, 'Title')

        # Test данные
        test = registry_record.find('Test')
        if test is not None:
            record['Дата'] = self._get_text(test, 'Date')
            record['№ протокола'] = self._get_text(test, 'ProtocolNumber')
            record['Программа'] = self._get_text(test, 'LearnProgramTitle')
            record['Результат'] = 'Удовлетворительно' if test.get('isPassed', 'false').lower() == 'true' else 'Неудовлетворительно'
            record['ID программы'] = test.get('learnProgramId', '')

        # Регистрационный номер (если есть)
        record['Регистрационный номер'] = self._get_text(registry_record, 'InternalExamination')

        # Внешний ID
        outer_id = registry_record.get('outerId', '')
        record['Внешний ID'] = outer_id

        return record

    def _parse_educated_person(self, educated_person: ET.Element) -> Dict[str, str]:
        """Парсинг одной записи EducatedPerson (из ответа GetEducatedPersonXML)"""
        record = {}

        # Прямые поля из EducatedPerson
        record['Фамилия'] = self._get_text(educated_person, 'LastName')
        record['Имя'] = self._get_text(educated_person, 'FirstName')
        record['Отчество'] = self._get_text(educated_person, 'MiddleName')
        record['СНИЛС'] = self._get_text(educated_person, 'Snils')
        record['Должность'] = self._get_text(educated_person, 'Position')
        record['ИНН Заказчика'] = self._get_text(educated_person, 'EmployerInn')
        record['Наименование ЮЛ Заказчика'] = self._get_text(educated_person, 'EmployerTitle')
        record['ID программы'] = self._get_text(educated_person, 'LearnProgramId')
        record['Дата'] = self._get_text(educated_person, 'Date')
        record['№ протокола'] = self._get_text(educated_person, 'ProtocolNumber')
        record['Результат'] = 'Удовлетворительно' if self._get_text(educated_person, 'IsPassed').lower() == 'true' else 'Неудовлетворительно'
        record['SetId'] = self._get_text(educated_person, 'SetId')

        return record

    def _get_text(self, parent: ET.Element, tag: str) -> str:
        """Получение текста элемента"""
        element = parent.find(tag)
        return element.text.strip() if element is not None and element.text else ''