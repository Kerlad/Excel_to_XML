"""
Альтернативный модуль работы с API Минтруда на базе pycurl.
PycURL - обёртка вокруг libcurl, максимальная совместимость с корпоративными прокси.
"""
import os
import io
import time
import json
import zipfile
import logging
import xml.etree.ElementTree as ET
from datetime import datetime
from urllib.parse import urlparse
from pycurl import Curl, READFUNCTION, POSTFIELDSIZE, HTTPHEADER, WRITEFUNCTION, POST, URL, VERBOSE, SSL_VERIFYPEER, SSL_VERIFYHOST, PROXY, PROXYUSERPWD, CUSTOMREQUEST

try:
    import pycurl
except ImportError:
    pycurl = None

logger = logging.getLogger(__name__)

API_URL = "https://edu.rosmintrud.ru/api/set/push"
GET_URL = "https://edu.rosmintrud.ru/api/GetEducatedPersonXML"

USER_AGENT = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'


def _setup_curl(proxy_settings=None, verbose=False):
    """
    Настройка объекта Curl с прокси.
    """
    c = Curl()
    c.setopt(SSL_VERIFYPEER, 0)
    c.setopt(SSL_VERIFYHOST, 0)
    c.setopt(USERAGENT, USER_AGENT)
    c.setopt(NOSIGNAL, 1)
    c.setopt(TIMEOUT, 60)
    
    if verbose:
        c.setopt(VERBOSE, 1)

    mode = proxy_settings.get("mode", "off") if proxy_settings else "off"

    if mode != "off":
        proxy_url = None
        proxy_auth = ""

        if mode == "auto":
            try:
                import winreg
                reg_path = r"Software\Microsoft\Windows\CurrentVersion\Internet Settings"
                with winreg.OpenKey(winreg.HKEY_CURRENT_USER, reg_path) as key:
                    proxy_enabled, _ = winreg.QueryValueEx(key, "ProxyEnable")
                    if proxy_enabled:
                        proxy_server, _ = winreg.QueryValueEx(key, "ProxyServer")
                        if proxy_server:
                            if "=" in proxy_server:
                                for part in proxy_server.split(";"):
                                    if part.startswith("https=") or part.startswith("http="):
                                        addr = part.split("=", 1)[1]
                                        if addr and not addr.startswith(("http://", "https://")):
                                            scheme = part.split("=")[0]
                                            addr = f"{scheme}://{addr}"
                                        proxy_url = addr
                                        break
                            else:
                                proxy_url = proxy_server if proxy_server.startswith(("http://", "https://")) else f"http://{proxy_server}"
            except Exception as e:
                logger.debug(f"Ошибка чтения реестра: {e}")
        elif mode == "manual":
            proxy_url = proxy_settings.get("url", "").strip()
            if not proxy_url.startswith(("http://", "https://")):
                proxy_url = f"http://{proxy_url}"
            
            username = proxy_settings.get("username", "").strip()
            password = proxy_settings.get("password", "").strip()
            if username and password:
                proxy_auth = f"{username}:{password}"

        if proxy_url:
            c.setopt(PROXY, proxy_url)
            if proxy_auth:
                c.setopt(PROXYUSERPWD, proxy_auth)
            logger.info(f"PycURL прокси: {proxy_url}")

    return c


def push_xml(api_key, xml_file_path, xsd_path=None, proxy_settings=None):
    """
    Отправка XML файла на сервер Минтруда через pycurl.
    """
    if pycurl is None:
        return {"success": False, "error": "Модуль pycurl не установлен. Установите: pip install pycurl"}

    if not api_key or len(api_key) != 32:
        return {"success": False, "error": "API-ключ должен содержать 32 символа"}

    if not os.path.exists(xml_file_path):
        return {"success": False, "error": "Файл XML не найден"}

    try:
        # Читаем XML данных
        with open(xml_file_path, 'rb') as f:
            data_xml_content = f.read()

        # Формируем Request.xml
        request_xml = f'''<?xml version="1.0" encoding="UTF-8"?>
<Request>
    <ApiKey>{api_key}</ApiKey>
    <NeedSend>false</NeedSend>
</Request>'''.encode('utf-8')

        # Создаём .olot архив (ZIP с Data.xml внутри)
        olot_buffer = io.BytesIO()
        with zipfile.ZipFile(olot_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.writestr('Data.xml', data_xml_content)
        olot_data = olot_buffer.getvalue()

        # Формируем multipart/form-data вручную
        boundary = '----WebKitFormBoundary7MA4YWxkTrZu0gW'
        
        body = io.BytesIO()
        
        # Добавляем Request.xml
        body.write(f'--{boundary}\r\n'.encode('utf-8'))
        body.write(b'Content-Disposition: form-data; name="xml"; filename="Request.xml"\r\n')
        body.write(b'Content-Type: text/xml\r\n\r\n')
        body.write(request_xml)
        body.write(b'\r\n')
        
        # Добавляем data.olot
        body.write(f'--{boundary}\r\n'.encode('utf-8'))
        body.write(b'Content-Disposition: form-data; name="olot"; filename="data.olot"\r\n')
        body.write(b'Content-Type: application/octet-stream\r\n\r\n')
        body.write(olot_data)
        body.write(b'\r\n')
        body.write(f'--{boundary}--\r\n'.encode('utf-8'))
        
        body_bytes = body.getvalue()

        # Настраиваем curl
        c = _setup_curl(proxy_settings)
        
        response_buffer = io.BytesIO()
        c.setopt(WRITEFUNCTION, response_buffer.write)
        
        headers = [
            f'Content-Type: multipart/form-data; boundary={boundary}',
            f'Content-Length: {len(body_bytes)}',
        ]
        c.setopt(HTTPHEADER, headers)
        c.setopt(URL, API_URL)
        c.setopt(POST, 1)
        c.setopt(POSTFIELDS, body_bytes)

        logger.info(f"Отправка XML на {API_URL} (pycurl)")
        c.perform()
        
        status_code = c.getinfo(HTTP_CODE)
        response_text = response_buffer.getvalue().decode('utf-8')
        c.close()

        if status_code == 200:
            try:
                root = ET.fromstring(response_text)
                set_id_elem = root.find('.//SetId')
                if set_id_elem is not None and set_id_elem.text:
                    return {
                        "success": True,
                        "set_id": set_id_elem.text.strip(),
                        "message": "Данные загружены на сервер"
                    }
            except ET.ParseError:
                pass
            
            return {
                "success": True,
                "set_id": "Не удалось извлечь SetId",
                "message": "Данные загружены, но ответ имеет неожиданный формат"
            }
        else:
            # Сохраняем ошибку
            try:
                error_log_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "log")
                os.makedirs(error_log_path, exist_ok=True)
                error_log_file = os.path.join(error_log_path, "error_response.txt")
                with open(error_log_file, 'w', encoding='utf-8-sig') as f:
                    f.write(f"HTTP Error: {status_code}\n")
                    f.write(f"URL: {API_URL}\n")
                    f.write(f"Response: {response_text}\n")
            except:
                pass
            
            return {"success": False, "error": f"HTTP {status_code}: {response_text[:200]}"}

    except pycurl.error as e:
        logger.error(f"PycURL ошибка: {e}")
        return {"success": False, "error": f"PycURL ошибка: {e}"}
    except Exception as e:
        logger.error(f"Критическая ошибка pycurl: {e}")
        return {"success": False, "error": f"Ошибка: {e}"}


def get_by_set_id(api_key, set_id, page_size=5000, proxy_settings=None):
    """
    Получение данных по SetId через pycurl.
    """
    if pycurl is None:
        return {"success": False, "error": "Модуль pycurl не установлен"}

    if not api_key or len(api_key) != 32:
        return {"success": False, "error": "API-ключ должен содержать 32 символа"}

    if not set_id:
        return {"success": False, "error": "SetId не указан"}

    # Формируем XML запрос
    xml_content = f'''<?xml version="1.0" encoding="UTF-8"?>
<Request>
    <ApiKey>{api_key}</ApiKey>
    <SetId>{set_id}</SetId>
    <PageSize>{page_size}</PageSize>
    <PageNo>1</PageNo>
</Request>'''.encode('utf-8')

    all_records = []
    page_no = 1

    while True:
        result = _fetch_page_pycurl(xml_content, f"стр. {page_no}", page_size, proxy_settings)
        if not result:
            break
        
        all_records.extend(result.get("records", []))
        
        if not result.get("has_more", False):
            break
        
        page_no += 1
        time.sleep(0.5)

    return {"success": True, "records": all_records}


def get_by_snils(api_key, snils, page_size=100, proxy_settings=None):
    """
    Получение данных по СНИЛС через pycurl.
    """
    if pycurl is None:
        return {"success": False, "error": "Модуль pycurl не установлен"}

    if not api_key or len(api_key) != 32:
        return {"success": False, "error": "API-ключ должен содержать 32 символа"}

    if not snils:
        return {"success": False, "error": "СНИЛС не указан"}

    # Формируем XML запрос
    xml_content = f'''<?xml version="1.0" encoding="UTF-8"?>
<Request>
    <ApiKey>{api_key}</ApiKey>
    <Snils>{snils}</Snils>
    <PageSize>{page_size}</PageSize>
    <PageNo>1</PageNo>
</Request>'''.encode('utf-8')

    all_records = []
    page_no = 1

    while True:
        result = _fetch_page_pycurl(xml_content, f"стр. {page_no}", page_size, proxy_settings)
        if not result:
            break
        
        all_records.extend(result.get("records", []))
        
        if not result.get("has_more", False):
            break
        
        page_no += 1
        time.sleep(0.5)

    return {"success": True, "records": all_records}


def _fetch_page_pycurl(xml_content, page_label="", page_size=100, proxy_settings=None):
    """
    Выполнение одного POST-запроса к GetEducatedPersonXML через pycurl.
    """
    if pycurl is None:
        return None

    try:
        # Формируем multipart/form-data
        boundary = '----WebKitFormBoundary7MA4YWxkTrZu0gW'
        
        body = io.BytesIO()
        body.write(f'--{boundary}\r\n'.encode('utf-8'))
        body.write(b'Content-Disposition: form-data; name="file"; filename="request.xml"\r\n')
        body.write(b'Content-Type: text/xml\r\n\r\n')
        body.write(xml_content if isinstance(xml_content, bytes) else xml_content.encode('utf-8'))
        body.write(b'\r\n')
        body.write(f'--{boundary}--\r\n'.encode('utf-8'))
        
        body_bytes = body.getvalue()

        # Настраиваем curl
        c = _setup_curl(proxy_settings)
        
        response_buffer = io.BytesIO()
        c.setopt(WRITEFUNCTION, response_buffer.write)
        
        headers = [
            f'Content-Type: multipart/form-data; boundary={boundary}',
            f'Content-Length: {len(body_bytes)}',
        ]
        c.setopt(HTTPHEADER, headers)
        c.setopt(URL, GET_URL)
        c.setopt(POST, 1)
        c.setopt(POSTFIELDS, body_bytes)

        logger.info(f"Запрос {page_label} через pycurl")
        c.perform()
        
        status_code = c.getinfo(HTTP_CODE)
        response_text = response_buffer.getvalue().decode('utf-8')
        c.close()

        if status_code != 200:
            logger.error(f"Ошибка HTTP {status_code} при запросе {page_label}")
            return None

        # Проверка на <Error> в теле
        if "<Error>" in response_text:
            try:
                root = ET.fromstring(response_text)
                if root.tag == "Error":
                    sc = root.find("StatusCode")
                    msg = root.find("Message")
                    sc_text = sc.text if sc is not None else ""
                    msg_text = msg.text if msg is not None else ""
                    logger.error(f"Логическая ошибка: {sc_text} - {msg_text}")
                    return None
            except ET.ParseError:
                pass

        # Проверка на наличие записей
        if "<RegistryRecord" not in response_text:
            return {"records": [], "has_more": False}

        # Парсинг записей
        records = _parse_registry_records(response_text)

        return {"records": records, "has_more": len(records) == page_size}

    except Exception as e:
        logger.error(f"Критическая ошибка запроса {page_label} (pycurl): {e}")
        return None


def _parse_registry_records(response_text):
    """Парсинг RegistryRecord из ответа."""
    records = []
    try:
        root = ET.fromstring(response_text)
        for record in root.findall('.//RegistryRecord'):
            rec = {}
            rec['baseNo'] = record.get('baseNo', '')
            rec['internalExamination'] = record.get('internalExamination', '')
            rec['setId'] = record.get('setId', '')
            rec['baseDateCreated'] = record.get('baseDateCreated', '')
            rec['outerId'] = record.get('outerId', '')

            worker = record.find('Worker')
            if worker is not None:
                rec['LastName'] = _tag_text(worker, 'LastName')
                rec['FirstName'] = _tag_text(worker, 'FirstName')
                rec['MiddleName'] = _tag_text(worker, 'MiddleName')
                rec['Snils'] = _tag_text(worker, 'Snils')
                rec['Position'] = _tag_text(worker, 'Position')

            test = record.find('Test')
            if test is not None:
                rec['learnProgramId'] = _tag_text(test, 'learnProgramId')
                rec['LearnProgramTitle'] = _tag_text(test, 'LearnProgramTitle')
                rec['ProtocolNumber'] = _tag_text(test, 'ProtocolNumber')
                rec['Date'] = _tag_text(test, 'Date')
                rec['Result'] = _tag_text(test, 'Result')

            records.append(rec)
    except ET.ParseError as e:
        logger.error(f"Ошибка парсинга XML (pycurl): {e}")

    return records


def _tag_text(parent, tag):
    """Получение текста элемента."""
    elem = parent.find(tag)
    return elem.text.strip() if elem is not None and elem.text else ""


def export_records_to_xlsx(records, file_path):
    """Экспорт записей в XLSX файл."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return False, "Не установлен модуль openpyxl"

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Записи"

        headers = [
            "Номер записи в реестре (baseNo)",
            "Фамилия (LastName)",
            "Имя (FirstName)",
            "Отчество (MiddleName)",
            "СНИЛС (Snils)",
            "Номер программы (learnProgramId)",
            "Название программы (LearnProgramTitle)",
            "Номер протокола (ProtocolNumber)",
            "Дата (Date)"
        ]
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4169E1", end_color="4169E1", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        def _format_date(date_val):
            """Конвертация даты из YYYY-MM-DD или YYYY-MM-DDTHH:MM:SS в ДД.ММ.ГГГГ."""
            if not date_val:
                return ''
            try:
                from datetime import datetime
                if 'T' in date_val:
                    dt = datetime.fromisoformat(date_val.replace('Z', '+00:00'))
                else:
                    dt = datetime.strptime(date_val[:10], "%Y-%m-%d")
                return dt.strftime("%d.%m.%Y")
            except (ValueError, TypeError):
                return date_val

        for rec in records:
            row = [
                rec.get('baseNo', ''),
                rec.get('LastName', ''),
                rec.get('FirstName', ''),
                rec.get('MiddleName', ''),
                rec.get('Snils', ''),
                rec.get('learnProgramId', ''),
                rec.get('LearnProgramTitle', ''),
                rec.get('ProtocolNumber', ''),
                _format_date(rec.get('Date', ''))
            ]
            ws.append(row)

        for col_idx, col in enumerate(ws.columns, 1):
            max_len = 0
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

        wb.save(file_path)
        return True, f"Сохранено {len(records)} записей\n{file_path}"

    except Exception as e:
        return False, f"Ошибка сохранения файла: {e}"
