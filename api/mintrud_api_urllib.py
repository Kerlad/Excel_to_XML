"""
Альтернативный модуль работы с API Минтруда на базе urllib.
Используется как резервный вариант при проблемах с requests через прокси.
"""
# ============================================================================
# TLS VERIFICATION SETTINGS
# ============================================================================
# Импортируем из proxy_manager - значение переопределяется чекбоксом в интерфейсе
from utils.proxy_manager import ENABLE_TLS_VERIFY
# ============================================================================

import os
import io
import time
import base64
import json
import zipfile
import logging
import xml.etree.ElementTree as ET
from datetime import datetime
from urllib.request import Request, urlopen, ProxyHandler, HTTPBasicAuthHandler, HTTPPasswordMgrWithDefaultRealm, build_opener, install_opener
from urllib.error import URLError, HTTPError
from urllib.parse import urlparse

logger = logging.getLogger(__name__)

API_URL = "https://edu.rosmintrud.ru/api/set/push"
GET_URL = "https://edu.rosmintrud.ru/api/GetEducatedPersonXML"

USER_AGENT = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'


def _setup_proxy_handler(proxy_settings):
    """
    Настройка обработчика прокси для urllib.
    Возвращает opener с настроенным прокси или None.
    """
    mode = proxy_settings.get("mode", "off") if proxy_settings else "off"

    if mode == "off":
        return None

    proxy_url = None
    if mode == "auto":
        # Автоопределение из Windows
        try:
            import winreg
            reg_path = r"Software\Microsoft\Windows\CurrentVersion\Internet Settings"
            with winreg.OpenKey(winreg.HKEY_CURRENT_USER, reg_path) as key:
                proxy_enabled, _ = winreg.QueryValueEx(key, "ProxyEnable")
                if proxy_enabled:
                    proxy_server, _ = winreg.QueryValueEx(key, "ProxyServer")
                    if proxy_server:
                        if "=" in proxy_server:
                            for part in proxy_server.split(";"):
                                if part.startswith("https=") or part.startswith("http="):
                                    addr = part.split("=", 1)[1]
                                    if addr and not addr.startswith(("http://", "https://")):
                                        scheme = part.split("=")[0]
                                        addr = f"{scheme}://{addr}"
                                    proxy_url = addr
                                    break
                        else:
                            proxy_url = proxy_server if proxy_server.startswith(("http://", "https://")) else f"http://{proxy_server}"
        except Exception as e:
            logger.debug(f"Ошибка чтения реестра: {e}")
    elif mode == "manual":
        proxy_url = proxy_settings.get("url", "").strip()
        if not proxy_url.startswith(("http://", "https://")):
            proxy_url = f"http://{proxy_url}"

    if not proxy_url:
        return None

    # Парсинг URL для извлечения credentials
    parsed = urlparse(proxy_url)
    proxy_host = f"{parsed.hostname}:{parsed.port}" if parsed.port else parsed.hostname

    # Создаем менеджер паролей
    password_mgr = HTTPPasswordMgrWithDefaultRealm()
    
    # Добавляем credentials если есть
    username = proxy_settings.get("username", "").strip() if mode == "manual" else (parsed.username or "")
    password = proxy_settings.get("password", "").strip() if mode == "manual" else (parsed.password or "")

    if username and password:
        # Для HTTPS через прокси нужно добавить прокси в password_mgr
        password_mgr.add_password(
            realm=None,
            uri=proxy_host,
            user=username,
            passwd=password
        )
        # Также добавляем для HTTP
        password_mgr.add_password(
            realm=None,
            uri=f"http://{proxy_host}",
            user=username,
            passwd=password
        )
        logger.info(f"Прокси с авторизацией: {proxy_host}")
    else:
        logger.info(f"Прокси без авторизации: {proxy_host}")

    # Создаем handlers
    proxy_handler = ProxyHandler({
        'http': proxy_url,
        'https': proxy_url,
    })
    
    auth_handler = HTTPBasicAuthHandler(password_mgr)
    
    # Создаем opener с SSL контекстом
    import ssl
    
    # TLS verification: use verified context for production, unverified for dev/testing
    if ENABLE_TLS_VERIFY:
        # Production: use default SSL context with system certificates
        context = ssl.create_default_context()
    else:
        # Development: ignore SSL certificate verification (INSECURE!)
        # WARNING: This exposes you to MITM attacks!
        context = ssl._create_unverified_context()
    
    opener = build_opener(proxy_handler, auth_handler)
    # Добавляем HTTPS handler с настроенным SSL контекстом
    from urllib.request import HTTPSHandler
    https_handler = HTTPSHandler(context=context)
    opener = build_opener(proxy_handler, auth_handler, https_handler)
    
    return opener


def push_xml(api_key, xml_file_path, xsd_path=None, proxy_settings=None):
    """
    Отправка XML файла на сервер Минтруда через urllib.
    """
    if not api_key or len(api_key) != 32:
        return {"success": False, "error": "API-ключ должен содержать 32 символа"}

    if not os.path.exists(xml_file_path):
        return {"success": False, "error": "Файл XML не найден"}

    try:
        # Читаем XML данных
        with open(xml_file_path, 'rb') as f:
            data_xml_content = f.read()

        # Формируем Request.xml
        request_xml = f'''<?xml version="1.0" encoding="UTF-8"?>
<Request>
    <ApiKey>{api_key}</ApiKey>
    <NeedSend>false</NeedSend>
</Request>'''.encode('utf-8')

        # Создаём .olot архив (ZIP с Data.xml внутри)
        olot_buffer = io.BytesIO()
        with zipfile.ZipFile(olot_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.writestr('Data.xml', data_xml_content)
        olot_data = olot_buffer.getvalue()

        # Формируем multipart/form-data вручную
        boundary = '----WebKitFormBoundary7MA4YWxkTrZu0gW'
        
        body = io.BytesIO()
        
        # Добавляем Request.xml
        body.write(f'--{boundary}\r\n'.encode('utf-8'))
        body.write(b'Content-Disposition: form-data; name="xml"; filename="Request.xml"\r\n')
        body.write(b'Content-Type: text/xml\r\n\r\n')
        body.write(request_xml)
        body.write(b'\r\n')
        
        # Добавляем data.olot
        body.write(f'--{boundary}\r\n'.encode('utf-8'))
        body.write(b'Content-Disposition: form-data; name="olot"; filename="data.olot"\r\n')
        body.write(b'Content-Type: application/octet-stream\r\n\r\n')
        body.write(olot_data)
        body.write(b'\r\n')
        body.write(f'--{boundary}--\r\n'.encode('utf-8'))
        
        body_bytes = body.getvalue()

        # Создаем запрос
        req = Request(API_URL, data=body_bytes)
        req.add_header('Content-Type', f'multipart/form-data; boundary={boundary}')
        req.add_header('User-Agent', USER_AGENT)

        # Настраиваем прокси
        opener = _setup_proxy_handler(proxy_settings)
        if opener:
            install_opener(opener)

        # Отправляем запрос
        logger.info(f"Отправка XML на {API_URL} (urllib)")
        with urlopen(req, timeout=60) as response:
            response_text = response.read().decode('utf-8')

        if response.status == 200:
            # Парсим ответ
            try:
                root = ET.fromstring(response_text)
                set_id_elem = root.find('.//SetId')
                if set_id_elem is not None and set_id_elem.text:
                    return {
                        "success": True,
                        "set_id": set_id_elem.text.strip(),
                        "message": "Данные загружены на сервер"
                    }
            except ET.ParseError:
                pass
            
            return {
                "success": True,
                "set_id": "Не удалось извлечь SetId",
                "message": "Данные загружены, но ответ имеет неожиданный формат"
            }
        else:
            return {"success": False, "error": f"HTTP {response.status}"}

    except HTTPError as e:
        error_text = e.read().decode('utf-8', errors='ignore') if hasattr(e, 'read') else ""
        logger.error(f"HTTP ошибка {e.code}: {error_text}")
        
        # Сохраняем ошибку
        try:
            error_log_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "log")
            os.makedirs(error_log_path, exist_ok=True)
            error_log_file = os.path.join(error_log_path, "error_response.txt")
            with open(error_log_file, 'w', encoding='utf-8-sig') as f:
                f.write(f"HTTP Error: {e.code}\n")
                f.write(f"URL: {API_URL}\n")
                f.write(f"Response: {error_text}\n")
        except:
            pass
        
        return {"success": False, "error": f"HTTP {e.code}: {error_text[:200]}"}
    except URLError as e:
        logger.error(f"URL ошибка: {e.reason}")
        return {"success": False, "error": f"Ошибка подключения: {e.reason}"}
    except Exception as e:
        logger.error(f"Критическая ошибка: {e}")
        return {"success": False, "error": f"Ошибка: {e}"}


def get_by_set_id(api_key, set_id, page_size=5000, proxy_settings=None):
    """
    Получение данных по SetId через urllib.
    """
    if not api_key or len(api_key) != 32:
        return {"success": False, "error": "API-ключ должен содержать 32 символа"}

    if not set_id:
        return {"success": False, "error": "SetId не указан"}

    # Формируем XML запрос
    xml_content = f'''<?xml version="1.0" encoding="UTF-8"?>
<Request>
    <ApiKey>{api_key}</ApiKey>
    <SetId>{set_id}</SetId>
    <PageSize>{page_size}</PageSize>
    <PageNo>1</PageNo>
</Request>'''.encode('utf-8')

    all_records = []
    page_no = 1

    while True:
        result = _fetch_page_urllib(xml_content, f"стр. {page_no}", page_size, proxy_settings)
        if not result:
            break
        
        all_records.extend(result.get("records", []))
        
        if not result.get("has_more", False):
            break
        
        page_no += 1
        time.sleep(0.5)

    return {"success": True, "records": all_records}


def get_by_snils(api_key, snils, page_size=100, proxy_settings=None):
    """
    Получение данных по СНИЛС через urllib.
    """
    if not api_key or len(api_key) != 32:
        return {"success": False, "error": "API-ключ должен содержать 32 символа"}

    if not snils:
        return {"success": False, "error": "СНИЛС не указан"}

    # Формируем XML запрос
    xml_content = f'''<?xml version="1.0" encoding="UTF-8"?>
<Request>
    <ApiKey>{api_key}</ApiKey>
    <Snils>{snils}</Snils>
    <PageSize>{page_size}</PageSize>
    <PageNo>1</PageNo>
</Request>'''.encode('utf-8')

    all_records = []
    page_no = 1

    while True:
        result = _fetch_page_urllib(xml_content, f"стр. {page_no}", page_size, proxy_settings)
        if not result:
            break
        
        all_records.extend(result.get("records", []))
        
        if not result.get("has_more", False):
            break
        
        page_no += 1
        time.sleep(0.5)

    return {"success": True, "records": all_records}


def _fetch_page_urllib(xml_content, page_label="", page_size=100, proxy_settings=None):
    """
    Выполнение одного POST-запроса к GetEducatedPersonXML через urllib.
    """
    try:
        # Формируем multipart/form-data
        boundary = '----WebKitFormBoundary7MA4YWxkTrZu0gW'
        
        body = io.BytesIO()
        body.write(f'--{boundary}\r\n'.encode('utf-8'))
        body.write(b'Content-Disposition: form-data; name="file"; filename="request.xml"\r\n')
        body.write(b'Content-Type: text/xml\r\n\r\n')
        body.write(xml_content if isinstance(xml_content, bytes) else xml_content.encode('utf-8'))
        body.write(b'\r\n')
        body.write(f'--{boundary}--\r\n'.encode('utf-8'))
        
        body_bytes = body.getvalue()

        # Создаем запрос
        req = Request(GET_URL, data=body_bytes)
        req.add_header('Content-Type', f'multipart/form-data; boundary={boundary}')
        req.add_header('User-Agent', USER_AGENT)

        # Настраиваем прокси
        opener = _setup_proxy_handler(proxy_settings)
        if opener:
            install_opener(opener)

        # Отправляем запрос
        logger.info(f"Запрос {page_label} через urllib")
        with urlopen(req, timeout=60) as response:
            response_text = response.read().decode('utf-8')

        if response.status != 200:
            logger.error(f"Ошибка HTTP {response.status} при запросе {page_label}")
            return None

        # Проверка на <Error> в теле
        if "<Error>" in response_text:
            try:
                root = ET.fromstring(response_text)
                if root.tag == "Error":
                    sc = root.find("StatusCode")
                    msg = root.find("Message")
                    sc_text = sc.text if sc is not None else ""
                    msg_text = msg.text if msg is not None else ""
                    logger.error(f"Логическая ошибка: {sc_text} - {msg_text}")
                    return None
            except ET.ParseError:
                pass

        # Проверка на наличие записей
        if "<RegistryRecord" not in response_text:
            return {"records": [], "has_more": False}

        # Парсинг записей
        records = _parse_registry_records(response_text)

        return {"records": records, "has_more": len(records) == page_size}

    except Exception as e:
        logger.error(f"Критическая ошибка запроса {page_label}: {e}")
        return None


def _parse_registry_records(response_text):
    """Парсинг RegistryRecord из ответа."""
    records = []
    try:
        root = ET.fromstring(response_text)
        for record in root.findall('.//RegistryRecord'):
            rec = {}
            # Атрибуты RegistryRecord
            rec['baseNo'] = record.get('baseNo', '')
            rec['internalExamination'] = record.get('internalExamination', '')
            rec['setId'] = record.get('setId', '')
            rec['baseDateCreated'] = record.get('baseDateCreated', '')
            rec['outerId'] = record.get('outerId', '')

            # Worker
            worker = record.find('Worker')
            if worker is not None:
                rec['LastName'] = _tag_text(worker, 'LastName')
                rec['FirstName'] = _tag_text(worker, 'FirstName')
                rec['MiddleName'] = _tag_text(worker, 'MiddleName')
                rec['Snils'] = _tag_text(worker, 'Snils')
                rec['Position'] = _tag_text(worker, 'Position')

            # Test
            test = record.find('Test')
            if test is not None:
                rec['learnProgramId'] = _tag_text(test, 'learnProgramId')
                rec['LearnProgramTitle'] = _tag_text(test, 'LearnProgramTitle')
                rec['ProtocolNumber'] = _tag_text(test, 'ProtocolNumber')
                rec['Date'] = _tag_text(test, 'Date')
                rec['Result'] = _tag_text(test, 'Result')

            records.append(rec)
    except ET.ParseError as e:
        logger.error(f"Ошибка парсинга XML: {e}")

    return records


def _tag_text(parent, tag):
    """Получение текста элемента."""
    elem = parent.find(tag)
    return elem.text.strip() if elem is not None and elem.text else ""


def export_records_to_xlsx(records, file_path):
    """
    Экспорт записей в XLSX файл.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return False, "Не установлен модуль openpyxl"

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Записи"

        # Заголовки
        headers = [
            "Номер записи в реестре (baseNo)",
            "Фамилия (LastName)",
            "Имя (FirstName)",
            "Отчество (MiddleName)",
            "СНИЛС (Snils)",
            "Номер программы (learnProgramId)",
            "Название программы (LearnProgramTitle)",
            "Номер протокола (ProtocolNumber)",
            "Дата (Date)"
        ]
        ws.append(headers)

        # Стилизация заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4169E1", end_color="4169E1", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Данные
        def _format_date(date_val):
            """Конвертация даты из YYYY-MM-DD или YYYY-MM-DDTHH:MM:SS в ДД.ММ.ГГГГ."""
            if not date_val:
                return ''
            try:
                from datetime import datetime
                if 'T' in date_val:
                    dt = datetime.fromisoformat(date_val.replace('Z', '+00:00'))
                else:
                    dt = datetime.strptime(date_val[:10], "%Y-%m-%d")
                return dt.strftime("%d.%m.%Y")
            except (ValueError, TypeError):
                return date_val

        for rec in records:
            row = [
                rec.get('baseNo', ''),
                rec.get('LastName', ''),
                rec.get('FirstName', ''),
                rec.get('MiddleName', ''),
                rec.get('Snils', ''),
                rec.get('learnProgramId', ''),
                rec.get('LearnProgramTitle', ''),
                rec.get('ProtocolNumber', ''),
                _format_date(rec.get('Date', ''))
            ]
            ws.append(row)

        # Автоширина столбцов
        for col_idx, col in enumerate(ws.columns, 1):
            max_len = 0
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

        wb.save(file_path)
        return True, f"Сохранено {len(records)} записей\n{file_path}"

    except Exception as e:
        return False, f"Ошибка сохранения файла: {e}"
