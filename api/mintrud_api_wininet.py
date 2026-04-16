"""
Альтернативный модуль работы с API Минтруда на базе WinINET API Windows.
Прямое использование системных функций Windows через ctypes для максимальной совместимости.
"""
import os
import io
import time
import json
import zipfile
import logging
import xml.etree.ElementTree as ET
from datetime import datetime
import ctypes
from ctypes import wintypes

logger = logging.getLogger(__name__)

API_URL = "https://edu.rosmintrud.ru/api/set/push"
GET_URL = "https://edu.rosmintrud.ru/api/GetEducatedPersonXML"

USER_AGENT = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'

# WinINET константы
INTERNET_OPEN_TYPE_PRECONFIG = 0
INTERNET_OPEN_TYPE_PROXY = 3
INTERNET_SERVICE_HTTP = 3
INTERNET_FLAG_RELOAD = 0x80000000
INTERNET_FLAG_NO_CACHE_WRITE = 0x04000000
INTERNET_FLAG_SECURE = 0x00800000
INTERNET_FLAG_IGNORE_CERT_CN_INVALID = 0x00001000
INTERNET_FLAG_IGNORE_CERT_DATE_INVALID = 0x00002000
HTTP_ADDREQ_FLAG_REPLACE = 0x80000000
HTTP_ADDREQ_FLAG_ADD = 0x20000000

# WinINET функции
try:
    wininet = ctypes.windll.wininet
except:
    wininet = None


def _wininet_http_request(url, method, headers, body, proxy_settings=None):
    """
    Выполнение HTTP-запроса через WinINET API.
    """
    if wininet is None:
        return None, "WinINET API недоступен"

    hInternet = None
    hConnect = None
    hRequest = None

    try:
        # Открываем сессию
        mode = proxy_settings.get("mode", "off") if proxy_settings else "off"
        
        if mode == "off":
            hInternet = wininet.InternetOpenW(
                USER_AGENT,
                INTERNET_OPEN_TYPE_PRECONFIG,
                None, None, 0
            )
        else:
            proxy_url = None
            if mode == "auto":
                # Автоопределение из Windows
                try:
                    import winreg
                    reg_path = r"Software\Microsoft\Windows\CurrentVersion\Internet Settings"
                    with winreg.OpenKey(winreg.HKEY_CURRENT_USER, reg_path) as key:
                        proxy_enabled, _ = winreg.QueryValueEx(key, "ProxyEnable")
                        if proxy_enabled:
                            proxy_server, _ = winreg.QueryValueEx(key, "ProxyServer")
                            if proxy_server:
                                if "=" in proxy_server:
                                    for part in proxy_server.split(";"):
                                        if part.startswith("https=") or part.startswith("http="):
                                            addr = part.split("=", 1)[1]
                                            if addr:
                                                proxy_url = addr
                                                break
                                else:
                                    proxy_url = proxy_server
                except Exception as e:
                    logger.debug(f"Ошибка чтения реестра: {e}")
            elif mode == "manual":
                proxy_url = proxy_settings.get("url", "").strip()
            
            if proxy_url:
                hInternet = wininet.InternetOpenW(
                    USER_AGENT,
                    INTERNET_OPEN_TYPE_PROXY,
                    proxy_url, None, 0
                )
            else:
                hInternet = wininet.InternetOpenW(
                    USER_AGENT,
                    INTERNET_OPEN_TYPE_PRECONFIG,
                    None, None, 0
                )

        if not hInternet:
            return None, f"Ошибка InternetOpenW: {ctypes.WinError()}"

        # Парсим URL
        from urllib.parse import urlparse
        parsed = urlparse(url)
        hostname = parsed.hostname
        port = parsed.port if parsed.port else (443 if parsed.scheme == 'https' else 80)
        path = parsed.path if parsed.path else '/'
        if parsed.query:
            path += f'?{parsed.query}'

        # Открываем соединение
        hConnect = wininet.InternetConnectW(
            hInternet,
            hostname,
            port,
            None, None,
            INTERNET_SERVICE_HTTP,
            0, 0
        )

        if not hConnect:
            return None, f"Ошибка InternetConnectW: {ctypes.WinError()}"

        # Открываем запрос
        flags = INTERNET_FLAG_RELOAD | INTERNET_FLAG_NO_CACHE_WRITE
        if parsed.scheme == 'https':
            flags |= INTERNET_FLAG_SECURE | INTERNET_FLAG_IGNORE_CERT_CN_INVALID | INTERNET_FLAG_IGNORE_CERT_DATE_INVALID

        hRequest = wininet.HttpOpenRequestW(
            hConnect,
            method,
            path,
            None, None,
            None,
            flags, 0
        )

        if not hRequest:
            return None, f"Ошибка HttpOpenRequestW: {ctypes.WinError()}"

        # Добавляем заголовки
        for key, value in headers.items():
            header_str = f"{key}: {value}"
            wininet.HttpAddRequestHeadersW(
                hRequest,
                header_str,
                len(header_str),
                HTTP_ADDREQ_FLAG_REPLACE | HTTP_ADDREQ_FLAG_ADD
            )

        # Отправляем запрос
        body_bytes = body.encode('utf-8') if isinstance(body, str) else body
        body_len = len(body_bytes)

        result = wininet.HttpSendRequestW(
            hRequest,
            None, 0,
            body_bytes, body_len
        )

        if not result:
            return None, f"Ошибка HttpSendRequestW: {ctypes.WinError()}"

        # Читаем ответ
        status_code = wintypes.DWORD()
        status_code_size = wintypes.DWORD(4)
        
        wininet.HttpQueryInfoW(
            hRequest,
            19,  # HTTP_QUERY_STATUS_CODE
            ctypes.byref(status_code),
            ctypes.byref(status_code_size),
            None
        )

        # Читаем тело ответа
        response_data = b""
        buffer = ctypes.create_string_buffer(4096)
        bytes_read = wintypes.DWORD()

        while True:
            if not wininet.InternetReadFile(hRequest, buffer, 4096, ctypes.byref(bytes_read)):
                break
            if bytes_read.value == 0:
                break
            response_data += buffer.raw[:bytes_read.value]

        return status_code.value, response_data.decode('utf-8')

    except Exception as e:
        logger.error(f"WinINET ошибка: {e}")
        return None, str(e)
    finally:
        # Закрываем хендлы
        if hRequest:
            wininet.InternetCloseHandle(hRequest)
        if hConnect:
            wininet.InternetCloseHandle(hConnect)
        if hInternet:
            wininet.InternetCloseHandle(hInternet)


def push_xml(api_key, xml_file_path, xsd_path=None, proxy_settings=None):
    """
    Отправка XML файла на сервер Минтруда через WinINET.
    """
    if not api_key or len(api_key) != 32:
        return {"success": False, "error": "API-ключ должен содержать 32 символа"}

    if not os.path.exists(xml_file_path):
        return {"success": False, "error": "Файл XML не найден"}

    try:
        # Читаем XML данных
        with open(xml_file_path, 'rb') as f:
            data_xml_content = f.read()

        # Формируем Request.xml
        request_xml = f'''<?xml version="1.0" encoding="UTF-8"?>
<Request>
    <ApiKey>{api_key}</ApiKey>
    <NeedSend>false</NeedSend>
</Request>'''.encode('utf-8')

        # Создаём .olot архив (ZIP с Data.xml внутри)
        olot_buffer = io.BytesIO()
        with zipfile.ZipFile(olot_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.writestr('Data.xml', data_xml_content)
        olot_data = olot_buffer.getvalue()

        # Формируем multipart/form-data
        boundary = '----WebKitFormBoundary7MA4YWxkTrZu0gW'
        
        body = io.BytesIO()
        
        body.write(f'--{boundary}\r\n'.encode('utf-8'))
        body.write(b'Content-Disposition: form-data; name="xml"; filename="Request.xml"\r\n')
        body.write(b'Content-Type: text/xml\r\n\r\n')
        body.write(request_xml)
        body.write(b'\r\n')
        
        body.write(f'--{boundary}\r\n'.encode('utf-8'))
        body.write(b'Content-Disposition: form-data; name="olot"; filename="data.olot"\r\n')
        body.write(b'Content-Type: application/octet-stream\r\n\r\n')
        body.write(olot_data)
        body.write(b'\r\n')
        body.write(f'--{boundary}--\r\n'.encode('utf-8'))
        
        body_bytes = body.getvalue()

        headers = {
            'User-Agent': USER_AGENT,
            'Content-Type': f'multipart/form-data; boundary={boundary}',
        }

        logger.info(f"Отправка XML на {API_URL} (WinINET)")

        status_code, response_text = _wininet_http_request(
            API_URL, "POST", headers, body_bytes, proxy_settings
        )

        if status_code is None:
            return {"success": False, "error": response_text}

        if status_code == 200:
            try:
                root = ET.fromstring(response_text)
                set_id_elem = root.find('.//SetId')
                if set_id_elem is not None and set_id_elem.text:
                    return {
                        "success": True,
                        "set_id": set_id_elem.text.strip(),
                        "message": "Данные загружены на сервер"
                    }
            except ET.ParseError:
                pass
            
            return {
                "success": True,
                "set_id": "Не удалось извлечь SetId",
                "message": "Данные загружены, но ответ имеет неожиданный формат"
            }
        else:
            # Сохраняем ошибку
            try:
                error_log_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "log")
                os.makedirs(error_log_path, exist_ok=True)
                error_log_file = os.path.join(error_log_path, "error_response.txt")
                with open(error_log_file, 'w', encoding='utf-8-sig') as f:
                    f.write(f"HTTP Error: {status_code}\n")
                    f.write(f"URL: {API_URL}\n")
                    f.write(f"Response: {response_text}\n")
            except:
                pass
            
            return {"success": False, "error": f"HTTP {status_code}: {response_text[:200]}"}

    except Exception as e:
        logger.error(f"Критическая ошибка WinINET: {e}")
        return {"success": False, "error": f"Ошибка: {e}"}


def get_by_set_id(api_key, set_id, page_size=5000, proxy_settings=None):
    """
    Получение данных по SetId через WinINET.
    """
    if not api_key or len(api_key) != 32:
        return {"success": False, "error": "API-ключ должен содержать 32 символа"}

    if not set_id:
        return {"success": False, "error": "SetId не указан"}

    # Формируем XML запрос
    xml_content = f'''<?xml version="1.0" encoding="UTF-8"?>
<Request>
    <ApiKey>{api_key}</ApiKey>
    <SetId>{set_id}</SetId>
    <PageSize>{page_size}</PageSize>
    <PageNo>1</PageNo>
</Request>'''.encode('utf-8')

    all_records = []
    page_no = 1

    while True:
        result = _fetch_page_wininet(xml_content, f"стр. {page_no}", page_size, proxy_settings)
        if not result:
            break
        
        all_records.extend(result.get("records", []))
        
        if not result.get("has_more", False):
            break
        
        page_no += 1
        time.sleep(0.5)

    return {"success": True, "records": all_records}


def get_by_snils(api_key, snils, page_size=100, proxy_settings=None):
    """
    Получение данных по СНИЛС через WinINET.
    """
    if not api_key or len(api_key) != 32:
        return {"success": False, "error": "API-ключ должен содержать 32 символа"}

    if not snils:
        return {"success": False, "error": "СНИЛС не указан"}

    # Формируем XML запрос
    xml_content = f'''<?xml version="1.0" encoding="UTF-8"?>
<Request>
    <ApiKey>{api_key}</ApiKey>
    <Snils>{snils}</Snils>
    <PageSize>{page_size}</PageSize>
    <PageNo>1</PageNo>
</Request>'''.encode('utf-8')

    all_records = []
    page_no = 1

    while True:
        result = _fetch_page_wininet(xml_content, f"стр. {page_no}", page_size, proxy_settings)
        if not result:
            break
        
        all_records.extend(result.get("records", []))
        
        if not result.get("has_more", False):
            break
        
        page_no += 1
        time.sleep(0.5)

    return {"success": True, "records": all_records}


def _fetch_page_wininet(xml_content, page_label="", page_size=100, proxy_settings=None):
    """
    Выполнение одного POST-запроса к GetEducatedPersonXML через WinINET.
    """
    try:
        # Формируем multipart/form-data
        boundary = '----WebKitFormBoundary7MA4YWxkTrZu0gW'
        
        body = io.BytesIO()
        body.write(f'--{boundary}\r\n'.encode('utf-8'))
        body.write(b'Content-Disposition: form-data; name="file"; filename="request.xml"\r\n')
        body.write(b'Content-Type: text/xml\r\n\r\n')
        body.write(xml_content if isinstance(xml_content, bytes) else xml_content.encode('utf-8'))
        body.write(b'\r\n')
        body.write(f'--{boundary}--\r\n'.encode('utf-8'))
        
        body_bytes = body.getvalue()

        headers = {
            'User-Agent': USER_AGENT,
            'Content-Type': f'multipart/form-data; boundary={boundary}',
        }

        logger.info(f"Запрос {page_label} через WinINET")

        status_code, response_text = _wininet_http_request(
            GET_URL, "POST", headers, body_bytes, proxy_settings
        )

        if status_code is None:
            logger.error(f"Ошибка WinINET при запросе {page_label}: {response_text}")
            return None

        if status_code != 200:
            logger.error(f"Ошибка HTTP {status_code} при запросе {page_label}")
            return None

        # Проверка на <Error> в теле
        if "<Error>" in response_text:
            try:
                root = ET.fromstring(response_text)
                if root.tag == "Error":
                    sc = root.find("StatusCode")
                    msg = root.find("Message")
                    sc_text = sc.text if sc is not None else ""
                    msg_text = msg.text if msg is not None else ""
                    logger.error(f"Логическая ошибка: {sc_text} - {msg_text}")
                    return None
            except ET.ParseError:
                pass

        # Проверка на наличие записей
        if "<RegistryRecord" not in response_text:
            return {"records": [], "has_more": False}

        # Парсинг записей
        records = _parse_registry_records(response_text)

        return {"records": records, "has_more": len(records) == page_size}

    except Exception as e:
        logger.error(f"Критическая ошибка запроса {page_label} (WinINET): {e}")
        return None


def _parse_registry_records(response_text):
    """Парсинг RegistryRecord из ответа."""
    records = []
    try:
        root = ET.fromstring(response_text)
        for record in root.findall('.//RegistryRecord'):
            rec = {}
            rec['baseNo'] = record.get('baseNo', '')
            rec['internalExamination'] = record.get('internalExamination', '')
            rec['setId'] = record.get('setId', '')
            rec['baseDateCreated'] = record.get('baseDateCreated', '')
            rec['outerId'] = record.get('outerId', '')

            worker = record.find('Worker')
            if worker is not None:
                rec['LastName'] = _tag_text(worker, 'LastName')
                rec['FirstName'] = _tag_text(worker, 'FirstName')
                rec['MiddleName'] = _tag_text(worker, 'MiddleName')
                rec['Snils'] = _tag_text(worker, 'Snils')
                rec['Position'] = _tag_text(worker, 'Position')

            test = record.find('Test')
            if test is not None:
                rec['learnProgramId'] = _tag_text(test, 'learnProgramId')
                rec['LearnProgramTitle'] = _tag_text(test, 'LearnProgramTitle')
                rec['ProtocolNumber'] = _tag_text(test, 'ProtocolNumber')
                rec['Date'] = _tag_text(test, 'Date')
                rec['Result'] = _tag_text(test, 'Result')

            records.append(rec)
    except ET.ParseError as e:
        logger.error(f"Ошибка парсинга XML (WinINET): {e}")

    return records


def _tag_text(parent, tag):
    """Получение текста элемента."""
    elem = parent.find(tag)
    return elem.text.strip() if elem is not None and elem.text else ""


def export_records_to_xlsx(records, file_path):
    """Экспорт записей в XLSX файл."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return False, "Не установлен модуль openpyxl"

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Записи"

        headers = [
            "Номер записи в реестре (baseNo)",
            "Фамилия (LastName)",
            "Имя (FirstName)",
            "Отчество (MiddleName)",
            "СНИЛС (Snils)",
            "Номер программы (learnProgramId)",
            "Название программы (LearnProgramTitle)",
            "Номер протокола (ProtocolNumber)",
            "Дата (Date)"
        ]
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4169E1", end_color="4169E1", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for rec in records:
            row = [
                rec.get('baseNo', ''),
                rec.get('LastName', ''),
                rec.get('FirstName', ''),
                rec.get('MiddleName', ''),
                rec.get('Snils', ''),
                rec.get('learnProgramId', ''),
                rec.get('LearnProgramTitle', ''),
                rec.get('ProtocolNumber', ''),
                rec.get('Date', '')
            ]
            ws.append(row)

        for col_idx, col in enumerate(ws.columns, 1):
            max_len = 0
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

        wb.save(file_path)
        return True, f"Сохранено {len(records)} записей\n{file_path}"

    except Exception as e:
        return False, f"Ошибка сохранения файла: {e}"
