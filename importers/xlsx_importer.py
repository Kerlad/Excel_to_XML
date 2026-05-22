"""
Модуль загрузки XLSX/XLS файлов со стримингом, прогрессом и отменой.
"""
import os
import logging
from datetime import datetime, timedelta

from utils.constants import VALID_PROGRAMS_SET as VALID_PROGRAMS
from utils.constants import MAX_XLSX_FILE_SIZE_MB, MAX_XLSX_ROWS
from utils.exceptions import FileTooLargeError, ImportLimitExceededError, ImportCancelledError

logger = logging.getLogger(__name__)

COLUMNS = [
    "Фамилия", "Имя", "Отчество", "СНИЛС", "Должность",
    "ИНН Заказчика", "Наименование ЮЛ Заказчика", "ИНН УЦ",
    "Наименование УЦ", "Результат", "№ программы", "Дата", "№ протокола"
]

FIELD_KEYS = [
    'last_name', 'first_name', 'middle_name', 'snils', 'position',
    'employer_inn', 'employer_title', 'tc_inn', 'tc_title',
    'result', 'program', 'date', 'protocol'
]


def format_snils(raw):
    """
    Приведение СНИЛС к формату '123-456-789 00'.
    Принимает любой вид: '12345678900', '123-456-78900', '123-456-789 00' и т.д.
    Возвращает отформатированную строку или None при невалидном вводе.
    Обрабатывает неразрывные пробелы (\xa0), табуляции и другие Unicode-пробелы.
    """
    clean = str(raw).strip()
    import unicodedata
    clean = ''.join(c for c in clean if unicodedata.category(c) != 'Zs')
    clean = clean.replace('-', '')
    if not clean.isdigit() or len(clean) != 11:
        return None
    return f"{clean[0:3]}-{clean[3:6]}-{clean[6:9]} {clean[9:11]}"


class FieldValidator:
    @staticmethod
    def validate_snils(value: str, row_num: int):
        formatted = format_snils(str(value))
        if formatted is None:
            masked = str(value)[:3] + '***' + str(value)[-2:] if len(str(value)) > 5 else '***'
            return {
                'row': row_num,
                'type': 'Ошибка',
                'field': 'СНИЛС',
                'message': f"СНИЛС должен содержать 11 цифр (введено: {masked})"
            }
        return None

    @staticmethod
    def validate_program(value: str, row_num: int):
        program_str = str(value).strip()
        programs = [p.strip() for p in program_str.rstrip(',').split(',') if p.strip()]
        invalid_programs = [p for p in programs if p not in VALID_PROGRAMS]
        if invalid_programs:
            return {
                'row': row_num,
                'type': 'Ошибка',
                'field': '№ программы',
                'message': f"Некорректный номер программы: {', '.join(invalid_programs)}"
            }
        return None

    @staticmethod
    def validate_result(value: str, row_num: int):
        result = str(value).strip()
        if result not in ['Удовлетворительно', 'Неудовлетворительно']:
            masked = result[:3] + '***' if len(result) > 3 else result
            return {
                'row': row_num,
                'type': 'Ошибка',
                'field': 'Результат',
                'message': f"Результат должен быть 'Удовлетворительно' или 'Неудовлетворительно' (введено: {masked})"
            }
        return None

    @staticmethod
    def validate_name(field_name: str, value, row_num: int):
        val = str(value).strip() if value is not None else ''
        if val and not val.replace(' ', '').replace('-', '').replace("'", '').isalpha():
            return {
                'row': row_num,
                'type': 'Ошибка',
                'field': field_name,
                'message': f"Поле '{field_name}' должно содержать только буквы"
            }
        return None

    @staticmethod
    def validate_date(value, row_num: int):
        if value is None:
            return {
                'row': row_num,
                'type': 'Ошибка',
                'field': 'Дата',
                'message': f"Дата некорректна"
            }
        if isinstance(value, datetime):
            return value.strftime('%d.%m.%Y')
        if isinstance(value, (int, float)):
            try:
                delta = datetime(1899, 12, 30) + timedelta(days=value)
                return delta.strftime('%d.%m.%Y')
            except (ValueError, OverflowError):
                return {
                    'row': row_num,
                    'type': 'Ошибка',
                    'field': 'Дата',
                    'message': f"Ошибка парсинга даты"
                }
        date_str = str(value).strip()
        if '.' in date_str or '-' in date_str:
            clean_date = date_str.replace('.', '').replace('-', '')
        else:
            clean_date = date_str
        if len(clean_date) == 8 and clean_date.isdigit():
            try:
                d = datetime.strptime(clean_date, "%d%m%Y")
                if d.date() > datetime.now().date():
                    return {
                        'row': row_num,
                        'type': 'Ошибка',
                        'field': 'Дата',
                        'message': f"Дата больше текущей"
                    }
                return f"{clean_date[:2]}.{clean_date[2:4]}.{clean_date[4:]}"
            except ValueError:
                return {
                    'row': row_num,
                    'type': 'Ошибка',
                    'field': 'Дата',
                    'message': f"Дата некорректна"
                }
        return {
            'row': row_num,
            'type': 'Ошибка',
            'field': 'Дата',
            'message': f"Дата некорректна"
        }

    @staticmethod
    def validate_required(col: str, value, row_num: int):
        if value is None or str(value).strip() == '':
            return {
                'row': row_num,
                'type': 'Ошибка',
                'field': col,
                'message': f"Пустое обязательное поле '{col}'"
            }
        return None


def validate_row(row_dict, row_num):
    """
    Валидация строки из Excel.
    Возвращает (True, data) или (False, error_details).
    error_details — список словарей: {'row': int, 'type': str, 'field': str, 'message': str}
    """
    errors = []

    required_cols = ['СНИЛС']
    for col in required_cols:
        err = FieldValidator.validate_required(col, row_dict.get(col), row_num)
        if err:
            errors.append(err)

    if errors:
        return False, errors

    snils = format_snils(str(row_dict['СНИЛС']))
    err = FieldValidator.validate_snils(row_dict['СНИЛС'], row_num)
    if err:
        errors.append(err)

    program_str = str(row_dict['№ программы']).strip()
    programs = [p.strip() for p in program_str.rstrip(',').split(',') if p.strip()]
    err = FieldValidator.validate_program(row_dict['№ программы'], row_num)
    if err:
        errors.append(err)

    err = FieldValidator.validate_result(row_dict['Результат'], row_num)
    if err:
        errors.append(err)

    for field_name in ['Фамилия', 'Имя', 'Отчество']:
        err = FieldValidator.validate_name(field_name, row_dict.get(field_name), row_num)
        if err:
            errors.append(err)

    date_str = FieldValidator.validate_date(row_dict['Дата'], row_num)
    if isinstance(date_str, dict):
        errors.append(date_str)
        date_str = None

    if errors:
        return False, errors

    records = []
    result = str(row_dict['Результат']).strip()
    for prog in programs:
        def gv(key):
            v = row_dict.get(key)
            return str(v).strip() if v is not None else ''

        record = {
            'last_name': gv('Фамилия'),
            'first_name': gv('Имя'),
            'middle_name': gv('Отчество'),
            'snils': snils,
            'position': gv('Должность'),
            'employer_inn': gv('ИНН Заказчика'),
            'employer_title': gv('Наименование ЮЛ Заказчика'),
            'tc_inn': gv('ИНН УЦ'),
            'tc_title': gv('Наименование УЦ'),
            'result': result,
            'program': prog,
            'date': date_str,
            'protocol': gv('№ протокола'),
            'source_row': row_num
        }
        records.append(record)

    return True, records


def load_xlsx(file_path, password=None, progress_callback=None, cancel_check=None):
    """
    Загрузка XLSX файла со стримингом, прогрессом и поддержкой отмены.
    
    Аргументы:
        file_path — путь к файлу
        password — пароль для защищённого файла (необязательно)
        progress_callback(current: int, total: int) — вызывается каждые 100 строк
            total=0 если неизвестно
        cancel_check() -> bool — вызывается каждую строку; True = отмена
    
    Возвращает (records, error_details, error_rows_set, error_msg).
    При отмене records=None, error_msg=["Импорт отменён"].
    """
    if not file_path.lower().endswith('.xlsx'):
        return None, [], set(), ["Неподдерживаемый формат. Используйте .xlsx"]

    try:
        size_mb = os.path.getsize(file_path) / (1024 * 1024)
        logger.info("Loading XLSX: %s (%.1f MB)", os.path.basename(file_path), size_mb)
        if size_mb > MAX_XLSX_FILE_SIZE_MB:
            msg = f"Файл превышает лимит {MAX_XLSX_FILE_SIZE_MB} МБ ({size_mb:.1f} МБ)"
            logger.warning("File too large: %s", msg)
            return None, [], set(), [msg]
    except OSError as e:
        logger.error("File access error: %s", e)
        return None, [], set(), [f"Ошибка доступа к файлу: {e}"]

    try:
        from openpyxl import load_workbook
        return _load_xlsx_openpyxl(file_path, password, progress_callback, cancel_check)
    except ImportError as e:
        logger.error("Missing openpyxl: %s", e)
        return None, [], set(), [f"Не установлен модуль: {e}. pip install openpyxl"]
    except FileTooLargeError as e:
        logger.warning(str(e))
        return None, [], set(), [str(e)]
    except ImportLimitExceededError as e:
        logger.warning(str(e))
        return None, [], set(), [str(e)]
    except ImportCancelledError:
        logger.info("XLSX import cancelled by user: %s", file_path)
        return None, [], set(), ["Импорт отменён пользователем"]


def _load_xlsx_openpyxl(file_path, password=None, progress_callback=None, cancel_check=None):
    """Загрузка .xlsx через openpyxl в read_only режиме со стримингом."""
    from openpyxl import load_workbook

    wb = load_workbook(file_path, data_only=True, read_only=True)
    try:
        ws = wb.active
        logger.debug("Active sheet: %s", ws.title)

        records = []
        error_details = []
        error_rows_set = set()

        headers = None
        row_count = 0
        progress_interval = 100

        for row in ws.iter_rows(values_only=True):
            if cancel_check and cancel_check():
                logger.info("Cancel requested at row %d", row_count)
                raise ImportCancelledError(f"Cancelled at row {row_count}")

            row_count += 1

            if row_count == 1:
                headers = [str(c or '').strip() for c in row]
                required_fields = ['СНИЛС']
                missing_cols = [col for col in required_fields if col not in headers]
                if missing_cols:
                    error_msg = f"Отсутствуют обязательные столбцы: {', '.join(missing_cols)}"
                    logger.warning("Missing columns: %s", missing_cols)
                    return None, [{"row": 1, "type": "Ошибка", "field": "Заголовки", "message": error_msg}], {1}, error_msg
                continue

            if row_count > MAX_XLSX_ROWS:
                logger.warning("Row limit exceeded: %d > %d", row_count, MAX_XLSX_ROWS)
                raise ImportLimitExceededError(
                    f"Превышено максимальное количество строк: {MAX_XLSX_ROWS}"
                )

            row_dict = {}
            all_empty = True
            for c_idx, col_name in enumerate(headers):
                val = row[c_idx] if c_idx < len(row) else None
                row_dict[col_name] = val if val is not None else ''
                if val is not None and str(val).strip() != '':
                    all_empty = False

            if all_empty:
                continue

            is_valid, result = validate_row(row_dict, row_count - 1)
            if is_valid:
                records.extend(result)
            else:
                error_details.extend(result)
                error_rows_set.add(row_count - 1)

            if progress_callback and row_count % progress_interval == 0:
                progress_callback(row_count, 0)

    finally:
        wb.close()
        logger.debug("Workbook closed: %s", file_path)

    if progress_callback:
        progress_callback(row_count, 0)

    logger.info("XLSX import completed: %d rows, %d records, %d errors",
                row_count - 1, len(records), len(error_details))

    return records, error_details, error_rows_set, ""
