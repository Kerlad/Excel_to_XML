"""
Модуль экспорта отчёта об ошибках импорта в XLSX
"""
import os
from datetime import datetime


def export_error_report(error_details, duplicate_map, file_path):
    """
    Экспорт отчёта об ошибках в XLSX файл.
    
    Args:
        error_details: список словарей {'row': int, 'type': str, 'field': str, 'message': str}
        duplicate_map: словарь {snils_program_key: [строка1, строка2, ...]} — дубликаты
        file_path: путь для сохранения
    
    Returns:
        (success: bool, message: str)
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        return False, "Не установлен модуль openpyxl"

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчёт об ошибках"

        # Заголовки
        headers = ["Тип ошибки", "Номер строки", "Поле", "Описание ошибки"]
        ws.append(headers)

        # Стилизация заголовков
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4169E1", end_color="4169E1", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Группируем ошибки по строкам для дубликатов
        row_errors_map = {}  # {row_num: [error_details]}
        for err in error_details:
            row_num = err['row']
            if row_num not in row_errors_map:
                row_errors_map[row_num] = []
            row_errors_map[row_num].append(err)
        for row_num in sorted(row_errors_map.keys()):
            errors = row_errors_map[row_num]
            for err in errors:
                row_data = [
                    err.get('type', 'Ошибка'),
                    err['row'],
                    err.get('field', ''),
                    err.get('message', '')
                ]
                ws.append(row_data)

        # Добавляем дубликаты
        for key, row_labels in duplicate_map.items():
            rows_str = "; ".join(str(r) for r in row_labels)
            snils, program = key
            
            # Формируем описание с аналогичными строками
            desc = f"Дубликат: СНИЛС={snils}, Программа={program}"
            # Добавляем аналогичные строки
            if len(row_labels) > 1:
                desc += f"\nАналогичные строки: {rows_str}"
            
            row_data = [
                "Дубликат",
                rows_str,
                "СНИЛС + № программы",
                desc
            ]
            ws.append(row_data)

        # Автоширина столбцов
        for col_idx, col in enumerate(ws.columns, 1):
            max_len = 0
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

        # Добавляем информацию о генерации
        ws.append([])
        ws.append([])
        ws.append([f"Отчёт сгенерирован: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}"])
        ws.append([f"Всего ошибок: {len(error_details)}"])
        ws.append([f"Всего строк с дубликатами: {len(duplicate_map)}"])

        wb.save(file_path)
        return True, f"Отчёт сохранён: {file_path}"

    except Exception as e:
        return False, f"Ошибка сохранения отчёта: {e}"
